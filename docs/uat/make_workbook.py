#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_workbook.py -- builds Kuja_UAT_Test_Plan.xlsx (Kuja Marketplace tenant only).

Every test case is grounded in the real routes / API endpoints / button labels of
the Kuja Grant app and the standalone Kuja Trust app. Test-data references point at
files in docs/uat/testfiles/ (built by make_testfiles.py).

Run:  py -3 docs/uat/make_workbook.py
Out:  docs/uat/Kuja_UAT_Test_Plan.xlsx  (+ copy to Downloads/Latest Kuja files)
"""
import os, sys, shutil
try: sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception: pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Kuja_UAT_Test_Plan.xlsx")
DL  = r"C:\Users\IdirisLoyan\Downloads\Latest Kuja files"

NAVY   = "0F3D5C"; NAVY2 = "17608A"; LIGHT = "EAF2F8"; ZEBRA = "F5F9FC"
GREY   = "6B7280"; AMBER = "9A6A00"; RED = "8A2A2A"
WHITE  = "FFFFFF"
thin   = Side(style="thin", color="C9D6E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ["Test ID","Module","Scenario","Type","Pri","Login as","Preconditions",
        "Test data / files","Steps","Expected result",
        "Actual result","Status","Tester","Date","Notes / Defect ID"]
WIDTHS = [11,20,30,9,5,13,26,26,54,50,26,11,12,11,20]
AUTHORED = 10  # first 10 columns are pre-filled; the rest are for the tester

TYPE_FILL = {
    "Happy":   "E7F4E9", "Edge": "FDF3E0", "Negative": "FBE9E9",
    "Security":"EDE7F6", "Setup": "EAF2F8", "Perf": "E6F4F1",
    "i18n":    "FDF0F5", "A11y": "EAF7F2",
}

def _title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1,1,text)
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26

def add_case_sheet(wb, name, tabcolor, intro, rows):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tabcolor
    ncols = len(COLS)
    _title(ws, f"{name}   —   {intro}", ncols)
    # header
    hr = 2
    for j,h in enumerate(COLS, start=1):
        c = ws.cell(hr, j, h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY2)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[hr].height = 24
    # data
    for i,row in enumerate(rows):
        r = hr + 1 + i
        vals = list(row) + [""]*(len(COLS)-AUTHORED)
        for j,v in enumerate(vals, start=1):
            c = ws.cell(r, j, v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top",
                wrap_text=True, horizontal="center" if j in (4,5,12) else "left")
            c.font = Font(size=9)
        # type color
        tfill = TYPE_FILL.get(row[3], ZEBRA if i % 2 else WHITE)
        ws.cell(r,4).fill = PatternFill("solid", fgColor=tfill)
        # zebra for the rest
        base = ZEBRA if i % 2 else WHITE
        for j in range(1, len(COLS)+1):
            if j != 4:
                ws.cell(r,j).fill = PatternFill("solid", fgColor=base)
        ws.cell(r,1).font = Font(size=9, bold=True, color=NAVY)
    # widths
    for j,w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    # freeze + filter
    ws.freeze_panes = "C3"
    last = hr + len(rows)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(ncols)}{last}"
    # status dropdown
    dv = DataValidation(type="list",
        formula1='"Pass,Fail,Blocked,Not run,In progress"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"L3:L{last}")
    return ws

def add_freeform(wb, name, tabcolor, blocks):
    """blocks = list of (kind, payload). kind in header/sub/para/bullets/table/space."""
    ws = wb.create_sheet(name); ws.sheet_properties.tabColor = tabcolor
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    for col in "CDEFG": ws.column_dimensions[col].width = 26
    r = 1
    def put(row, col, text, **kw):
        c = ws.cell(row, col, text)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal=kw.get("h","left"), indent=kw.get("indent",0))
        c.font = Font(size=kw.get("size",10), bold=kw.get("bold",False),
                      color=kw.get("color","000000"), italic=kw.get("italic",False))
        if kw.get("fill"): c.fill = PatternFill("solid", fgColor=kw["fill"])
        return c
    for kind, payload in blocks:
        if kind == "title":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            put(r,2,payload,size=15,bold=True,color=WHITE,fill=NAVY); ws.row_dimensions[r].height=30; r+=2
        elif kind == "header":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            put(r,2,payload,size=12,bold=True,color=NAVY,fill=LIGHT); ws.row_dimensions[r].height=22; r+=1
        elif kind == "sub":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            put(r,2,payload,size=10,bold=True,color=NAVY2); r+=1
        elif kind == "para":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c=put(r,2,payload,size=10)
            ws.row_dimensions[r].height = max(16, 14*(1+len(payload)//95)); r+=1
        elif kind == "bullet":
            put(r,2,"•",h="center"); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
            c=put(r,3,payload,size=10); ws.row_dimensions[r].height=max(15,14*(1+len(payload)//80)); r+=1
        elif kind == "warn":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            put(r,2,payload,size=10,bold=True,color=RED,fill="FBE9E9")
            ws.row_dimensions[r].height=max(18,14*(1+len(payload)//90)); r+=1
        elif kind == "table":
            headers, data = payload
            for j,h in enumerate(headers):
                put(r,2+j,h,bold=True,color=WHITE,fill=NAVY2,h="left")
                ws.cell(r,2+j).border=BORDER
            r+=1
            for k,drow in enumerate(data):
                for j,v in enumerate(drow):
                    c=put(r,2+j,v,size=9,fill=ZEBRA if k%2 else WHITE); c.border=BORDER
                ws.row_dimensions[r].height=max(15,14*(1+len(str(drow[-1]))//40)); r+=1
        elif kind == "space":
            r += payload if isinstance(payload,int) else 1
    return ws

# ---------------------------------------------------------------------------
# DATA — populated below (each *_ROWS is a list of 10-tuples matching AUTHORED)
#   (Test ID, Module, Scenario, Type, Pri, Login as, Preconditions,
#    Test data/files, Steps, Expected result)
# ---------------------------------------------------------------------------
N = "NGO · fatima@amani.org"
NGO_ROWS = [
 ("NGO-001","Sign in","Sign in with valid NGO credentials","Happy","P1",N,
  "On a demo host or the prod URL; signed out","—",
  "1. Go to /login\n2. Enter email fatima@amani.org and password pass123\n3. Click 'Sign in'",
  "Lands on /dashboard showing the NGO (member) dashboard; name/org shown; no console errors."),
 ("NGO-002","Sign in","Wrong password is rejected","Negative","P1",N,
  "Signed out","—",
  "1. Go to /login\n2. Enter fatima@amani.org with a wrong password\n3. Click 'Sign in'",
  "Inline error shown; no session created; not redirected to /dashboard."),
 ("NGO-003","Sign in","Account lockout after repeated failures","Security","P2",N,
  "Signed out; a throwaway demo email","—",
  "1. Attempt sign-in with a wrong password 5 times within 5 minutes",
  "After 5 failures the account is locked ~15 minutes; a lockout message appears; correct password is refused until the window passes."),
 ("NGO-004","Sign in","Forced password change on first login","Security","P1",N,
  "Admin has just reset this user (must_change_password = true)","—",
  "1. Sign in\n2. Observe redirect to /change-password\n3. Set a new password (>= 10 chars, different from current)\n4. Continue",
  "All /api calls are blocked until the password is changed; after setting it, normal access resumes at /dashboard."),
 ("NGO-005","Localisation","Switch UI language to Arabic (RTL)","i18n","P2",N,
  "Signed in","—",
  "1. Open /settings (or the language switcher)\n2. Choose Arabic\n3. Navigate to /dashboard and /grants",
  "UI renders in Arabic; page direction flips to RTL (dir='rtl'); choice persists across reloads. (Login page stays LTR by design.)"),
 ("NGO-006","Org profile","View organisation profile","Happy","P2",N,
  "Signed in","—","1. Go to /organizations/profile",
  "Profile loads current org fields (name, country, year, budget, staff, website, mission, sectors) and a completion % bar."),
 ("NGO-007","Org profile","Edit and save profile","Happy","P1",N,
  "Signed in","—",
  "1. Edit name, country, year established, annual budget, staff count, website, mission\n2. Select sectors (e.g. Health, Nutrition)\n3. Click 'Save changes'",
  "Save succeeds (PUT /api/organizations/{id}); success toast; completion % increases to reflect filled fields."),
 ("NGO-008","Org profile","Completion meter reacts to missing fields","Edge","P3",N,
  "Signed in","—","1. Clear the 'mission' field\n2. Save\n3. Re-add it and save again",
  "Completion % drops then rises; it counts the 8 tracked fields (name/country/year/budget/staff/website/mission/sectors)."),
 ("NGO-009","Org profile","Export org data — JSON","Happy","P3",N,
  "Signed in","—","1. On the profile page click 'JSON'",
  "A JSON org-bundle downloads (/api/exports/org-bundle?format=json-download)."),
 ("NGO-010","Org profile","Export org data — ZIP with PDFs","Happy","P3",N,
  "Signed in","—","1. Click 'ZIP (with PDFs)'",
  "A .zip downloads containing the org bundle plus attached PDFs."),
 ("NGO-011","Browse grants","See open grants","Happy","P1",N,
  "Signed in; at least one published grant exists","—","1. Go to /grants ('Browse')",
  "Open grants render as cards with donor, deadline, sector; no draft/withdrawn grants appear to the NGO."),
 ("NGO-012","Browse grants","Search filters the list","Happy","P2",N,
  "Signed in","—","1. Type a keyword in the search box",
  "List filters live by title/donor/description; clearing the box restores the full list."),
 ("NGO-013","Browse grants","Sector filter chips + Clear filters","Happy","P2",N,
  "Signed in","—","1. Click a sector chip (e.g. Health)\n2. Click 'Clear filters'",
  "Only matching grants show; 'Clear filters' resets all facets."),
 ("NGO-014","Browse grants","Closing-in and country facets + sort","Edge","P3",N,
  "Signed in","—","1. Set 'Closing in' to 7d then 30d\n2. Pick a country chip\n3. Change Sort by Deadline/Funding/Recent",
  "Results filter by deadline window and country; the country facet self-hides when <= 1 country; sort reorders correctly."),
 ("NGO-015","Browse grants","Save a search and re-apply it","Edge","P3",N,
  "Signed in","—","1. Set some filters\n2. In the saved-searches bar, save the current search\n3. Reload and apply it",
  "The saved search persists (/api/saved-searches) and re-applies the same filters."),
 ("NGO-016","Browse grants","Watchlist star toggle","Happy","P2",N,
  "Signed in","—","1. Click the star on a grant card\n2. Confirm it appears in the watchlist\n3. Click the star again",
  "Star toggles on/off (POST /api/watchlist/toggle returns {starred}); watchlist reflects the change."),
 ("NGO-017","Grant detail","Express interest notifies the donor","Happy","P2",N,
  "Signed in; grant is open and not yet applied","—",
  "1. Open a grant /grants/{id}\n2. Click 'Express interest'",
  "Toast 'Donor notified you might apply.'; the donor receives an expression-of-interest signal."),
 ("NGO-018","Grant detail","Compare fit across grants (NGO only)","Edge","P3",N,
  "Signed in","—","1. On /grants toggle 'Compare fit'\n2. Select up to 4 grants\n3. Click 'Run comparison'",
  "A comparison dialog scores the org's fit for the selected grants (POST /api/grants/fit-compare)."),
 ("NGO-019","Grant detail","Ask a question in Grant Q&A","Happy","P3",N,
  "Signed in","—","1. Open a grant, go to the Q&A tab\n2. Post a question",
  "The question is recorded and visible; the donor can answer it later."),
 ("NGO-020","Apply","Start an application (create draft)","Happy","P1",N,
  "Signed in; grant is open; org has not applied yet","—",
  "1. On an open grant click 'Apply' (goes to /apply/{grantId})\n2. On step 1 click Next",
  "A draft application is created (POST /api/applications/); the header shows a 'Saved' stamp; wizard advances to Proposal."),
 ("NGO-021","Apply","Resume an existing draft (no duplicate)","Edge","P2",N,
  "Signed in; a draft already exists for this grant","—",
  "1. Click 'Apply' again for the same grant",
  "The existing draft resumes (the 409 'already applied' is handled); no second application is created."),
 ("NGO-022","Apply","Eligibility gate blocks progress","Negative","P1",N,
  "Signed in; in the apply wizard","—",
  "1. On the Eligibility step leave a REQUIRED requirement unchecked\n2. Click Next",
  "Progress is blocked with a toast; you cannot reach the Proposal step until all required items are checked."),
 ("NGO-023","Apply","Import eligibility evidence from profile","Happy","P3",N,
  "Signed in; org profile is filled","—",
  "1. On the Eligibility step click 'Import from profile'",
  "Evidence textareas auto-fill from the org profile."),
 ("NGO-024","Apply","AI 'Draft for me' fills an empty criterion","Happy","P2",N,
  "Signed in; on the Proposal step; a criterion is empty","—",
  "1. On an empty criterion click 'Draft for me'",
  "AI-generated text populates the field (POST /api/ai/draft-section); an AI telemetry entry is logged; wording is editable."),
 ("NGO-025","Apply","AI 'Strengthen against this criterion'","Happy","P2",N,
  "Signed in; a criterion already has text","—",
  "1. Click 'Strengthen against this criterion'",
  "Returns strengths/gaps/sharpened text and replaces the response (POST /api/ai/strengthen-section)."),
 ("NGO-026","Apply","AI Polish requires a minimum length","Edge","P3",N,
  "Signed in; on the Proposal step","—",
  "1. With < 30 chars, open More > Polish\n2. Then add > 30 chars and Polish again",
  "Polish is guarded below 30 chars; above it, a clarity-only rewrite is returned (POST /api/ai/polish-response)."),
 ("NGO-027","Apply","AI guidance + quality ring","Happy","P3",N,
  "Signed in","—","1. Open More > AI guidance on a criterion\n2. Click 'Apply Suggestions'",
  "Guidance and a quality score ring appear; 'Apply Suggestions' appends suggestions to the response."),
 ("NGO-028","Apply","Rubric live preview + word-count meter","Happy","P3",N,
  "Signed in; on the Proposal step","—",
  "1. Type into a criterion and watch the side panel and word count",
  "The rubric live preview updates per keystroke; the word count vs max_words changes colour as it nears the limit."),
 ("NGO-029","Apply","Upload required documents (each type)","Happy","P1",N,
  "Signed in; on the Documents step","01/02/03: registration cert, proposal PDF/DOCX, budget XLSX, logframe XLSX",
  "1. For each document requirement, drag-drop or click to upload the matching file (accept .pdf,.doc,.docx,.xls,.xlsx)",
  "Each file uploads (POST /api/documents/upload with doc_type) and shows an AI-review score; required chips clear."),
 ("NGO-030","Apply","Review step tallies completeness","Happy","P2",N,
  "Signed in; eligibility+proposal+docs done","—","1. Go to the Review step",
  "Shows Eligibility X/N, Proposal X/N, Documents X/N; a warning banner appears only if items are missing."),
 ("NGO-031","Apply","Submit is gated until complete","Negative","P1",N,
  "Signed in; one required doc OR one criterion is missing","—",
  "1. Try to click 'Submit Application'",
  "Submit is disabled while anything required is missing; if forced, the backend returns 400 'missing_criteria'."),
 ("NGO-032","Apply","Submit a complete application","Happy","P1",N,
  "Signed in; all required eligibility/criteria/docs done","03_grant_application/* set",
  "1. Complete every step\n2. Click 'Submit Application'",
  "Status becomes 'submitted'; an AI score ring and a 'What happens next' panel appear; reviewers may be auto-assigned."),
 ("NGO-033","Apply","Submit after the deadline is blocked","Negative","P2",N,
  "A grant whose deadline has passed","—","1. Try to submit an application to a closed-deadline grant",
  "Submission is refused with a 400 'deadline passed'."),
 ("NGO-034","Apply","Draft autosaves and survives offline","Edge","P2",N,
  "Signed in; in the apply wizard","—",
  "1. Edit a criterion\n2. Turn off the network\n3. Edit again and click Next\n4. Restore the network",
  "Edits queue in the offline outbox and drain on reconnect (last-write-wins); the 'Saved' stamp catches up; no data lost."),
 ("NGO-035","Lifecycle","Applications list + attention strip","Happy","P2",N,
  "Signed in; has drafts + submitted apps","—","1. Go to /applications\n2. Use the status filter chips",
  "Chips (draft/submitted/under_review/…); attention strip shows 'Continue N drafts' / 'N awaiting decision'."),
 ("NGO-036","Lifecycle","Download my applications CSV","Happy","P3",N,
  "Signed in","—","1. On /applications click 'Download CSV'",
  "A CSV of the org's applications downloads (/api/applications/my-applications.csv)."),
 ("NGO-037","Lifecycle","Withdraw a submitted application","Happy","P2",N,
  "Signed in; an application is 'submitted'","—",
  "1. Open the application\n2. Click 'Withdraw'\n3. Enter a reason and confirm",
  "Status becomes 'withdrawn' (POST /api/applications/{id}/withdraw); it leaves the active pipeline."),
 ("NGO-038","Lifecycle","Withdraw is unavailable from draft","Negative","P3",N,
  "Signed in; a draft application","—","1. Open a draft application and look for Withdraw",
  "No withdraw action (or a 400) — withdraw is only allowed from 'submitted'."),
 ("NGO-039","Lifecycle","Revision requested → edit and resubmit","Happy","P1",N,
  "Donor has set the app to 'revision_requested' (see DON tab)","—",
  "1. Open the application\n2. Read the 'Revision requested' banner + donor feedback\n3. Click 'Edit + resubmit'\n4. Make changes and resubmit",
  "The wizard reopens the same application; on resubmit the status returns to 'submitted'."),
 ("NGO-040","Lifecycle","Why-rejected explanation on a declined app","Happy","P2",N,
  "An application is 'declined'/'rejected'","—","1. Open the declined application",
  "A 'Why rejected' panel shows an AI, constructive explanation (GET /api/applications/{id}/explain-rejection)."),
 ("NGO-041","Lifecycle","Appeal a declined decision","Happy","P2",N,
  "An application is 'declined'/'rejected'","—",
  "1. Open the Appeal panel\n2. Enter a reason of >= 20 characters and submit",
  "Appeal is accepted (POST /api/applications/{id}/appeal); a reason < 20 chars is rejected; appealing twice is idempotent."),
 ("NGO-042","Lifecycle","Appeal blocked when not declined","Negative","P3",N,
  "An application is 'submitted' or 'awarded'","—","1. Try to appeal",
  "No appeal is allowed unless the application is declined/rejected."),
 ("NGO-043","Lifecycle","Acknowledge decision feedback + download PDF","Happy","P3",N,
  "A decided application","—","1. Open it, mark feedback viewed\n2. Click 'PDF'",
  "Feedback marked viewed (POST /feedback-viewed); an application PDF downloads (/api/applications/{id}.pdf)."),
 ("NGO-044","Compliance","Compliance dashboard shows deliverables & risk","Happy","P2",N,
  "Signed in; org has an awarded grant with deliverables","—","1. Go to /compliance",
  "Per-grant deliverables with due dates, risk dots, overdue badges, AI compliance score; draft grants excluded; summary tiles (At risk / Overdue / Avg compliance)."),
 ("NGO-045","Compliance","Create a report","Happy","P1",N,
  "Signed in; has an awarded grant","—","1. Go to /reports\n2. Create a report (grant, type, period, title)",
  "A draft report is created (POST /api/reports/)."),
 ("NGO-046","Compliance","Report attachment Upload button","Negative","P1",N,
  "Signed in; a draft report exists","03_grant_application/Amani_Detailed_Budget.pdf",
  "1. On /reports expand a draft report\n2. Click 'Upload' and choose the PDF",
  "SPEC: the file attaches to the report. KNOWN ISSUE to verify: POST /api/reports/{id}/attachments may 404 (route not implemented) — if so, raise a defect and use photo-evidence / apply-time upload instead."),
 ("NGO-047","Compliance","Photo-as-evidence upload + AI extraction","Happy","P2",N,
  "Signed in; a draft/revision report","04_compliance_reporting/field_photo_CHP_training.jpg (or .webp), receipt, attendance",
  "1. Open the report's AI tools\n2. 'Add photo evidence', choose photo_type (training/receipt/attendance/site_visit)\n3. Upload the image\n4. Click 'Extract details'",
  "Image attaches (POST /reports/{id}/photo-evidence, jpg/png/webp/gif, <= 5 MB) as report_evidence; AI vision extracts details."),
 ("NGO-048","Compliance","Photo evidence over 5 MB is rejected","Negative","P3",N,
  "Signed in; a draft report","Any image > 5 MB (or 99_edge_cases/oversized_20MB.pdf renamed)",
  "1. Attempt to add an oversized image as photo evidence",
  "Rejected client-side (5 MB cap) / backend rejects non-image or empty; a clear error is shown."),
 ("NGO-049","Compliance","Voice-to-report structures a transcript","Happy","P3",N,
  "Signed in; a draft/revision report","—",
  "1. Open the Voice report composer\n2. Speak/enter a transcript\n3. Generate",
  "The transcript (<= 12000 chars) is mapped onto the donor's reporting requirements (POST /reports/{id}/structure-from-voice); works only on draft/revision."),
 ("NGO-050","Compliance","Pre-submit check surfaces gaps","Happy","P2",N,
  "Signed in; a draft report with partial content","—","1. Run the pre-submit review/precheck",
  "Shows 'Ready to submit' or 'Address gaps first' with per-requirement scores; no status change (POST /reports/{id}/precheck)."),
 ("NGO-051","Compliance","Submit a report","Happy","P1",N,
  "Signed in; a complete draft report","04_compliance_reporting/Amani_Q1_Narrative_Progress_Report.docx + Financial xlsx",
  "1. Attach the narrative + financial files\n2. Click 'Submit now'",
  "Status becomes 'submitted'; AI analysis runs; the donor's report inbox receives it."),
 ("NGO-052","Compliance","Request a reporting deadline extension","Edge","P3",N,
  "Signed in; a report near/after due date","—","1. Open the deadline negotiator / request extension",
  "An extension request is logged (POST /reports/{id}/extension-request); the donor can approve/deny it."),
]
TA = "Kuja Trust app · demo workspace"
TRUST_ROWS = [
 ("TRU-001","In-app assessment","Open the capacity assessment area","Happy","P2",N,
  "Signed in to the Grant app","—","1. Go to /assessments",
  "The assessments list and a 'start assessment' wizard are available (this is the Grant app's own in-app assessment)."),
 ("TRU-002","In-app assessment","Start the wizard and pick a framework","Happy","P2",N,
  "Signed in","—","1. Start a new assessment\n2. Choose 'Kuja Standard' (26 items, 5 domains)",
  "The checklist for the chosen framework renders with domains/weights."),
 ("TRU-003","In-app assessment","Answer the checklist and see a score","Happy","P2",N,
  "Signed in; wizard open","—","1. Answer the checklist items\n2. Save",
  "A score is computed (_calculate_assessment_scores) and the org's assess_score / assess_date update."),
 ("TRU-004","In-app assessment","Attach evidence to an assessment","Happy","P2",N,
  "Signed in; an assessment exists","02_capacity_trust_evidence/Amani_Safeguarding_PSEA_Policy.pdf",
  "1. Upload a policy PDF as assessment evidence",
  "Document attaches (POST /api/assessments/{id}/documents; shared 16 MB / allowed-types rules apply)."),
 ("TRU-005","Trust Passport (in-app)","View the two-pillar Trust Profile","Happy","P2",N,
  "Signed in","—","1. Go to /trust",
  "A two-pillar profile shows Capacity and Due-Diligence components with statuses."),
 ("TRU-006","Trust Passport (in-app)","Publish a Passport","Happy","P2",N,
  "Signed in; profile has content","—","1. On /trust click Publish passport",
  "A passport is published and a public share slug is generated (POST /api/passport/publish)."),
 ("TRU-007","Trust Passport (in-app)","Public share page needs no login","Happy","P2","(anonymous)",
  "A passport has been published","—",
  "1. Copy the share link\n2. Open /trust/share/{slug} in a private/incognito window",
  "The passport renders with no sign-in; a revoked/expired slug returns 410 (gone)."),
 ("TRU-008","Trust Passport (in-app)","Verify a passport with a token","Happy","P2","(funder/anon)",
  "A verify link+token exists","—","1. Open /trust/verify/{slug}?t={token}",
  "Verified state (200) is shown; a missing token prompts for one; an invalid token returns 401."),
 ("TRU-009","Trust Passport (in-app)","Revoke then verify returns gone","Edge","P3",N,
  "A published passport","—","1. Revoke the passport\n2. Re-open the share/verify link",
  "After revocation the public links return 410 (gone)."),
 ("TRU-010","Grant↔Trust","Trust readiness is a soft nudge, not a gate","Edge","P1",N,
  "Signed in; org has NO passport","—",
  "1. Start/submit an application without a Trust profile/passport",
  "A soft CTA ('Open Trust') may appear, but application submission is NOT blocked — readiness returns ready:true; grant submit has no passport/trust gate."),
 ("TRU-011","Grant↔Trust","Seam diagnostics show the engine is LIVE","Setup","P2","Admin · admin@kuja.org",
  "Signed in as admin","—","1. Call GET /api/admin/trust-engine/status (admin diagnostics)",
  "For the Kuja tenant with the hand-off configured: mode='remote', remote_configured=true, remote_reachable=true. (If the shared secret / engine are unset it falls back to 'local' — the kill switch.)"),
 ("TRU-012","Grant↔Trust","Hand-off: Grant → Kuja Trust assessment (no 2nd login)","Happy","P1",N,
  "Signed in as NGO on /trust; hand-off configured","—",
  "1. On /trust click 'Complete your capacity assessment in Kuja Trust'\n2. Follow the redirect",
  "You land on the Assessment tab in Kuja Trust with NO second login (a short-lived signed link binds your org); a 'Return to your application' bar is shown. Your Trust org is created/linked on first arrival (external_ref=grant:<id>)."),
 ("TRU-013","Trust app — access","Enter the demo workspace (no credentials)","Happy","P1",TA,
  "On https://kuja-app-production.up.railway.app","—",
  "1. Click 'Explore the demo workspace'",
  "A demo session opens bound to a seeded org; no login required; the workspace loads."),
 ("TRU-014","Trust app — access","Workspace shows all six tabs","Happy","P2",TA,
  "In the demo workspace","—","1. Note the tab bar",
  "Tabs present: Overview, Assessment, Evidence room, Frameworks, Verifiers, Passport."),
 ("TRU-015","Trust app — assessment","Rate the 7 C4C domains (1–4 rubric)","Happy","P1",TA,
  "In the Assessment tab","—",
  "1. For each domain (General, Governance, Strategy, Accountability, People & safeguarding, Finance, Data & IT) pick a rating: High risk / Basic controls / Established / Exceeds standard\n2. Add a note",
  "Each domain records a rating; the capacity score band updates accordingly."),
 ("TRU-016","Trust app — assessment","'I'm not sure' writes no score","Edge","P2",TA,
  "In the Assessment tab","—","1. On a domain choose the 'I'm not sure' path",
  "No score is recorded for that domain; the profile is treated as incomplete for that area (by design)."),
 ("TRU-017","Trust app — assessment","Import a C4C workbook","Edge","P3",TA,
  "In the Assessment tab","A C4C-format .xlsx/.xlsm/.csv (team-provided)",
  "1. Use 'Import C4C workbook' and select the file",
  "Domain answers import from the workbook (POST /api/import/c4c)."),
 ("TRU-018","Trust app — assessment","Delegate a domain to a colleague","Edge","P3",TA,
  "In the Assessment tab","—","1. 'Delegate a domain' and generate the link",
  "A single-use link valid ~14 days is created for a colleague to complete that domain."),
 ("TRU-019","Trust app — evidence","Upload allowed evidence types","Happy","P1",TA,
  "In the Evidence room","02_.../Amani_HR_Policy.pdf; 04_.../field_photo_CHP_training.webp",
  "1. Upload a PDF and a WebP image (also try JPEG/PNG/DOCX)",
  "Files store successfully (allowed: PDF, JPEG, PNG, WebP, DOCX; <= 8 MB)."),
 ("TRU-020","Trust app — evidence","Oversized evidence is rejected (413)","Negative","P2",TA,
  "In the Evidence room","99_edge_cases/oversized_20MB.pdf",
  "1. Try to upload the 20 MB file",
  "Rejected with 413 'file_too_large' (8 MB cap); nothing is stored."),
 ("TRU-021","Trust app — evidence","Active-content PDF is refused (415)","Security","P1",TA,
  "In the Evidence room","99_edge_cases/pdf_with_active_javascript.pdf",
  "1. Try to upload the PDF that contains JavaScript/OpenAction",
  "Refused with 415 'rejected_by_screening'; bytes are NOT stored; the refusal is written to the audit trail (evidence.refused)."),
 ("TRU-022","Trust app — evidence","Legacy .doc and MIME mismatch refused","Negative","P2",TA,
  "In the Evidence room","99_edge_cases/not_really_a_pdf.pdf (text renamed .pdf)",
  "1. Upload a legacy binary .doc (if available)\n2. Upload not_really_a_pdf.pdf",
  "Legacy .doc is rejected deliberately; a declared-vs-detected MIME mismatch is rejected."),
 ("TRU-023","Trust app — evidence","Encrypted/no-body file is held or refused","Security","P3",TA,
  "In the Evidence room","99_edge_cases/password_protected.pdf",
  "1. Try to upload the password-protected PDF",
  "Encrypted/empty-body content is refused or quarantined (held_for_screening / rejected), never silently accepted."),
 ("TRU-024","Trust app — screening","Run sanctions/PEP/adverse-media screening","Happy","P1",TA,
  "In the workspace","—","1. Click 'Run screen'",
  "Screening runs (OpenSanctions sanctions/PEP, SAM.gov exclusions, AI adverse media); results show a 'configured' flag per source."),
 ("TRU-025","Trust app — screening","'Verified clear' never over self-declared rows","Security","P1",TA,
  "Screening has run","—",
  "1. Inspect the status wording on independent checks (sanctions/PEP/SAM/adverse media) vs self-declared rows",
  "Only independent checks can read 'verified clear'; self-declared rows never display 'Verified clear' (single StatusText renderer, assurance-aware)."),
 ("TRU-026","Trust app — screening","Incomplete profile shows honest status","Security","P1",TA,
  "Profile is incomplete","—","1. View the passport/status while some domains are unanswered",
  "Shows 'available checks clear — profile incomplete' and names the missing pieces — never a blanket 'verified clear'."),
 ("TRU-027","Trust app — screening","Overall status = worst pillar + score band","Edge","P2",TA,
  "Both pillars scored","—","1. Compare overall status to the two pillar statuses",
  "Overall is the WORST of the pillars, folded with the score band (< 40 flagged, < 70 review, >= 70 clear)."),
 ("TRU-028","Trust app — passport","Publish, share and verify with a signed receipt","Happy","P2",TA,
  "Assessment + screening complete","—",
  "1. Publish the passport\n2. Open the public share page\n3. Verify it and record a funder decision",
  "Public share renders (no token); verify shows a proof layer (VC url, did:web); a signed VC receipt downloads and re-verifies at POST /api/credentials/verify."),
 ("TRU-029","Trust app — i18n","Only complete languages are offered","i18n","P2",TA,
  "In the workspace / public share","—","1. Open the language selector on a public share page",
  "en, fr, ar, sw are offered; es and so are withdrawn (dictionaries incomplete) — they must NOT appear as options."),
 ("TRU-030","Trust app — i18n","Arabic renders right-to-left","i18n","P2",TA,
  "In the workspace / public share","—","1. Switch to Arabic on the share/verify page and the workspace",
  "Direction flips to RTL (dir='rtl') on public surfaces and the workspace; long-form workspace guidance may stay English."),
 ("TRU-031","Trust app — low-data","Low-data mode is server-enforced","Security","P2",TA,
  "In the workspace","—",
  "1. Turn on low-data mode\n2. Trigger an AI route (e.g. a full-profile review)",
  "AI routes return 409 'low_data_mode'; the mode can only be turned ON (monotonic); hiding UI is cosmetic — the server refuses regardless."),
 ("TRU-032","Trust app — admin","Pilot cross-org read is time-boxed and audited","Security","P3","Trust admin (provisioned)",
  "Platform-admin Trust access","—","1. As a Trust platform admin, read another org during the pilot window",
  "Cross-org READ is allowed until KUJA_ADMIN_FULL_READ_UNTIL and every read is audited; cross-org WRITE is never allowed."),
 ("TRU-033","Grant↔Trust","Return to grant + result reads back","Happy","P1",N,
  "Completed a hand-off assessment in Kuja Trust (TRU-012)","—",
  "1. In Kuja Trust click 'Return to your application'\n2. Land back on Grant /trust (?from=trust)",
  "You return to the Grant app; a 'Welcome back from Kuja Trust' banner shows and the capacity & due-diligence status now reflects the Trust assessment (the profile is served with source='trust')."),
 ("TRU-034","Grant↔Trust","Hand-off availability gate / kill switch","Edge","P2",N,
  "Signed in as NGO","—",
  "1. Call GET /api/trust/handoff/available\n2. (Ops) confirm the CTA hides if KUJA_TRUST_HANDOFF_SECRET is unset",
  "Returns available:true only when the hand-off is configured for the Kuja tenant; with the secret unset the CTA is hidden, /handoff is inert, and the app falls back to its own in-app assessment — non-destructive kill switch."),
]
D = "Donor · sarah@globalhealth.org"
DONOR_ROWS = [
 ("DON-001","Sign in","Donor signs in to the donor dashboard","Happy","P1",D,
  "Signed out","—","1. Sign in as sarah@globalhealth.org / pass123",
  "Lands on the donor (grantmaker) dashboard with portfolio tiles."),
 ("DON-002","Create grant","AI-extract a grant from a Call-for-Proposals PDF","Happy","P1",D,
  "Signed in","05_donor/GlobalHealthFund_Call_for_Proposals_MCH.pdf",
  "1. /grants > 'Create'\n2. Step 0: drop the CfP PDF (accept .pdf,.doc,.docx,.txt)",
  "A draft grant is created then the PDF is uploaded (POST /api/grants/{id}/upload-grant-doc, <= 16 MB); AI extracts reporting requirements/indicators and pre-fills the wizard."),
 ("DON-003","Create grant","Skip AI and/or use the brief prompt scaffold","Edge","P3",D,
  "In the create wizard","—","1. Click the 'skip' link on step 0\n2. Or type a 1–2 line brief and generate a scaffold",
  "Skipping goes straight to manual entry; the brief prompt returns an AI scaffold (POST /api/ai/donor-grant-copilot)."),
 ("DON-004","Create grant","Fill basic info","Happy","P1",D,
  "In the create wizard","—",
  "1. Step 1: enter title (required), description, total funding, currency (USD/EUR/GBP/KES/CHF), deadline, sectors, countries",
  "Values accept and persist; title is required."),
 ("DON-005","Create grant","Set eligibility requirements with weights","Happy","P2",D,
  "In the create wizard","—",
  "1. Step 2: enable categories (geographic/org_type/experience/budget/sector/registration), set weight sliders and details\n2. Add a custom requirement",
  "Requirements save with weights; custom requirement is added."),
 ("DON-006","Create grant","Evaluation criteria must total 100%","Negative","P1",D,
  "In the create wizard","—",
  "1. Step 3: add criteria with weights that do NOT sum to 100\n2. Try to proceed/publish",
  "The weight total shows red and the backend rejects criteria whose weights don't sum to ~100."),
 ("DON-007","Create grant","Design criteria with AI + templates","Edge","P3",D,
  "In the create wizard, step 3","—",
  "1. Click 'Design with AI'\n2. Apply a saved criteria template\n3. Save current criteria as a template",
  "AI suggests criteria; a template applies/saves (GET/POST /api/grants/criteria-templates)."),
 ("DON-008","Create grant","Required document requirements","Happy","P2",D,
  "In the create wizard, step 4","—",
  "1. Enable document types (financial/registration/audit/PSEA/project/budget/CV/strategic) and mark some required",
  "Document requirements save with the required flag."),
 ("DON-009","Create grant","Save as draft","Happy","P2",D,
  "In the create wizard, step 5","—","1. Click 'Save draft'",
  "Grant saves as draft (PUT /api/grants/{id}); a review warning lists any missing title/funding/deadline."),
 ("DON-010","Create grant","Publish a grant","Happy","P1",D,
  "A complete draft; donor is licensed","—","1. On step 5 click 'Publish'",
  "Draft → open; published_at set; saved-search alerts + 'grant.published' webhook + smart-match notifications fire; NGOs can now see it."),
 ("DON-011","Create grant","Publish is blocked without a licence","Negative","P1",D,
  "Donor org lacks a grant licence","—","1. Try to publish",
  "Publish returns 403 'license_required' and the global 'upgrade required' dialog appears instead of an error."),
 ("DON-012","Grant mgmt","Export grants CSV","Happy","P3",D,
  "Signed in","—","1. On /grants click 'Export CSV'",
  "A CSV of the donor's grants downloads (/api/exports/grants.csv)."),
 ("DON-013","Grant mgmt","Duplicate a grant","Edge","P3",D,
  "Signed in; owns a grant","—","1. On a grant detail click 'Duplicate'",
  "A new '(copy)' grant is created in draft with the deadline cleared (POST /api/grants/{id}/duplicate)."),
 ("DON-014","Grant mgmt","Withdraw a published grant (cascade)","Edge","P2",D,
  "Signed in; owns an open grant with applications","—",
  "1. On grant detail click 'Withdraw grant'\n2. Enter a reason and confirm",
  "Grant and every open application flip to 'withdrawn'; applicants are notified (POST /api/grants/{id}/withdraw)."),
 ("DON-015","Grant mgmt","Delete a draft grant only","Negative","P3",D,
  "Signed in; a draft and an open grant","—","1. Delete a draft grant\n2. Try to delete an open grant",
  "Draft deletes (optionally cascading draft apps); an open/published grant cannot be deleted."),
 ("DON-016","Review pipeline","Table vs Pipeline (kanban) toggle","Happy","P2",D,
  "Signed in; grant has applications","—","1. Go to /applications\n2. Toggle 'Table' / 'Pipeline'",
  "Both views render; kanban groups applications by status."),
 ("DON-017","Decisions","Move an application to a decision","Happy","P1",D,
  "Owns the grant; an application is submitted; donor licensed","—",
  "1. In kanban/table set status to under_review, then awarded or rejected",
  "Status updates (PATCH /api/applications/{id}/status); donor must own the grant and be licensed (else 403 license_required)."),
 ("DON-018","Decisions","Request a revision from the NGO","Happy","P1",D,
  "Owns the grant; an application is submitted","—",
  "1. Open the application\n2. 'Request revision' and enter feedback",
  "Status → 'revision_requested' with feedback (POST /request-revision); the NGO sees the banner (pairs with NGO-039)."),
 ("DON-019","Decisions","Request an additional document","Edge","P2",D,
  "Owns the grant; an application is submitted","—",
  "1. 'Request document' with a label and optional note",
  "A document request is recorded (label required; POST /request-document); the NGO is notified."),
 ("DON-020","Review pipeline","Shortlist star + bulk unstar + filter","Edge","P2",D,
  "Signed in; several applications","—",
  "1. Star two applications\n2. Filter by '★ Shortlisted'\n3. Select and 'Unstar N'",
  "Stars persist (POST /star); the shortlist filter shows only starred; bulk unstar clears them (POST /bulk-star)."),
 ("DON-021","Review pipeline","Side-by-side compare applications","Happy","P2",D,
  "Signed in; >= 2 applications on one grant","—",
  "1. Go to /applications/compare (or ?ids=1,2,3)\n2. Compare",
  "A criteria × applications matrix renders with Status / AI / Human per column (GET /api/applications/compare)."),
 ("DON-022","Decisions","AI vs human vs final scores are visible","Happy","P2",D,
  "An application has AI + human scores","—","1. Open the application; view the score tiles + breakdown",
  "AI / Human / Final rings and a per-criterion breakdown are shown."),
 ("DON-023","Decisions","Decision audit drawer is populated","Security","P2",D,
  "A decided application","—","1. Open the decision audit drawer",
  "A hash-chained timeline of status transitions and AI calls is shown."),
 ("DON-024","Decisions","Record a win/loss debrief","Edge","P3",D,
  "An awarded or rejected application","—",
  "1. Open the debrief panel\n2. Pick a reason code + notes and save",
  "Debrief saves (PUT /debrief) only on awarded/rejected; reason vocab comes from /decision-reasons."),
 ("DON-025","Review pipeline","Assign / auto-assign reviewers","Happy","P2",D,
  "A submitted application; a reviewer account exists","—",
  "1. Assign james@reviewer.org (or use auto-assign / suggest reviewers)",
  "A review is created (POST /api/reviews/), the application moves to under_review, and the NGO is notified."),
 ("DON-026","Appeals","Respond to an NGO appeal","Happy","P2",D,
  "An NGO filed an appeal (see NGO-041)","—",
  "1. Open /admin/appeals or the AppealPanel\n2. Resolve the appeal with an outcome",
  "The appeal resolves (POST /appeal/resolve); the NGO sees the outcome."),
 ("DON-027","Portfolio","Ask about my grantees (portfolio Q&A)","Happy","P2",D,
  "Signed in; has grantees/reports","—",
  "1. Go to /portfolio-qa\n2. Ask e.g. 'Which of my grantees look at-risk right now?'",
  "An answer with clickable citations to grants/applications/reports is returned (POST /api/donor/portfolio-qa), scoped to this donor's portfolio."),
 ("DON-028","Portfolio","Portfolio Q&A is donor-only","Negative","P2",N,
  "Signed in as an NGO","—","1. As an NGO, navigate to /portfolio-qa",
  "Access is refused with a 'Donor accounts only' message; no portfolio data is exposed."),
 ("DON-029","Portfolio","Donor dashboard tiles load","Happy","P3",D,
  "Signed in","—","1. Open the donor dashboard",
  "Tiles render (scorecard, repeat grantees, decision forecast, review pipeline, starred queue) without errors."),
 ("DON-030","Grantee reports","Accept a submitted report","Happy","P1",D,
  "An NGO submitted a report on the donor's grant (NGO-051)","—",
  "1. Open the donor report inbox on /reports\n2. Click 'Accept'",
  "Report is accepted (POST /api/reports/{id}/review {action:accept}); status updates; NGO notified."),
 ("DON-031","Grantee reports","Request a revision on a report","Happy","P2",D,
  "A submitted report","—","1. In the report inbox click 'Request revision' with a note",
  "Report → revision_requested with the note; the NGO can revise and resubmit."),
 ("DON-032","Grantee reports","Private reviewer notes stay hidden from NGO","Security","P1",D,
  "A report/application with reviewer private notes","—",
  "1. As donor/reviewer add private notes\n2. Sign in as the NGO and open the same item",
  "Private notes are never shown to the NGO."),
 ("DON-033","Exports","Application PDF + audit-folder ZIP","Happy","P3",D,
  "Signed in; owns a grant with applications","—",
  "1. Download an application PDF (/api/applications/{id}.pdf)\n2. Download the grant audit folder (/api/grants/{id}/audit-folder)",
  "Both download; the audit ZIP is available to donor/NGO-owner/admin but NOT to reviewers."),
 ("DON-034","Grant mgmt","Cannot edit a grant you don't own","Negative","P1",D,
  "A grant owned by a DIFFERENT donor exists","—","1. Attempt to PUT/edit another donor's grant",
  "The action is refused with 403 (donor_org_id ownership check)."),
]
R = "Reviewer · james@reviewer.org"
REVIEWER_ROWS = [
 ("REV-001","Queue","Reviewer signs in to the reviewer queue","Happy","P1",R,
  "Signed out","—","1. Sign in as james@reviewer.org / pass123\n2. Go to /reviews",
  "Reviewer dashboard/queue loads with 'Pending (n)' and 'Completed (n)' tabs."),
 ("REV-002","Queue","Pending rows show due-soon urgency","Happy","P2",R,
  "Has assigned reviews; one due < 24h","—","1. View the Pending tab",
  "Each row shows applicant, grant, status; a 'Due Nh' badge appears when < 24h remain."),
 ("REV-003","Scoring","Open a review from the queue","Happy","P1",R,
  "An application is assigned to this reviewer","—",
  "1. Click 'Start review' on a pending row (goes to /reviews/{id})",
  "The scoring screen opens for that application; status may flip to in_progress."),
 ("REV-004","Scoring","Score criteria and save a draft","Happy","P1",R,
  "In a review","05_donor/GlobalHealthFund_Scoring_Rubric.pdf (reference)",
  "1. Set a 0–100 score + comment per criterion\n2. Click 'Save draft'",
  "A weighted overall score is computed; the draft persists (PUT /api/reviews/{id}); status assigned → in_progress."),
 ("REV-005","Scoring","Submit / complete a review","Happy","P1",R,
  "A drafted review","—","1. Click Submit / complete",
  "Review completes (POST /api/reviews/{id}/complete); it moves to the Completed tab; the donor sees the score."),
 ("REV-006","AI tools","AI triage summary of an application","Happy","P2",R,
  "In a review","—","1. Open 'AI triage summary' / 'Summarize'",
  "A concise AI summary of the application is returned (POST /api/ai/summarize-application); telemetry logged."),
 ("REV-007","AI tools","AI auto-score the application","Happy","P2",R,
  "In a review","—","1. Click the AI auto-score action",
  "AI proposes scores across criteria (POST /api/ai/score-application) as suggestions the reviewer can accept or override."),
 ("REV-008","AI tools","AI suggest a rationale per criterion","Happy","P3",R,
  "In a review","—","1. On a criterion click 'suggest rationale'",
  "An AI rationale is proposed for that criterion (POST /api/ai/score-criterion)."),
 ("REV-009","AI tools","Extract evidence from the application","Edge","P3",R,
  "In a review","—","1. Click 'extract evidence'",
  "Key evidence snippets are extracted (POST /api/ai/extract-evidence)."),
 ("REV-010","AI tools","Accept AI score + divergence warning","Edge","P2",R,
  "In a review with AI scores present","—",
  "1. Click 'Accept AI score'\n2. Then manually set a human score > 25 points from AI",
  "'Accept AI score' snaps human inputs to the AI overall; a tip warns when a human score diverges from AI by > 25."),
 ("REV-011","Integrity","Private notes are reviewer-only","Security","P1",R,
  "In a review","—","1. Type private notes; they autosave on blur\n2. Confirm the NGO cannot see them (NGO login)",
  "Notes persist (PUT /api/reviews/{id} private_notes, <= 8000 chars) and are never visible to the NGO."),
 ("REV-012","Integrity","COI disclosure auto-recuses the reviewer","Security","P1",R,
  "Assigned a review where a conflict exists","—",
  "1. Open the COI banner\n2. Disclose a conflict (employer_overlap/prior_consulting/family/other) with a note",
  "The disclosure is recorded and written to the audit chain (review.coi_disclosed); admins are notified; the reviewer is auto-recused (review removed; app reverts to submitted if no other reviewer)."),
 ("REV-013","Integrity","Decline an assignment","Happy","P2",R,
  "An assigned review","—","1. Click 'Decline assignment', give a reason, confirm",
  "The assignment is declined (POST /api/reviews/{id}/decline)."),
 ("REV-014","Integrity","Snooze a review","Edge","P3",R,
  "An assigned review","—","1. Snooze for 3 / 7 / 14 days with a reason",
  "The review is snoozed (POST /api/reviews/{id}/snooze) and resurfaces after the chosen period."),
 ("REV-015","AI tools","Reviewer recommendation / compare matrix","Edge","P3",R,
  "Several pending reviews","—",
  "1. Select up to 5 rows\n2. Open 'Side-by-side matrix' / 'Compare'",
  "A ranked fund/clarify/decline recommendation with similarity alerts is shown (POST /api/ai/reviewer-recommendation)."),
 ("REV-016","Integrity","Reviewer cannot self-assign reviews","Negative","P1",R,
  "Signed in as reviewer","—","1. Attempt to create/assign a review to yourself (POST /api/reviews/)",
  "Refused — review creation is restricted to donor/admin; reviewers cannot self-assign."),
 ("REV-017","Integrity","Reviewer cannot open the audit-folder ZIP","Negative","P2",R,
  "Signed in as reviewer","—","1. Attempt GET /api/grants/{id}/audit-folder",
  "Access is refused for reviewers (audit folder is donor/NGO-owner/admin only)."),
 ("REV-018","Queue","Reviewer stat strips populate","Happy","P3",R,
  "Has completed reviews","—","1. View caseload / turnaround / calibration strips",
  "Stats load from the reviewer endpoints (/api/reviews/my-caseload, my-turnaround, my-calibration, …)."),
]
A = "Admin · admin@kuja.org"
ADMIN_ROWS = [
 ("ADM-001","Sign in","Admin signs in to the operator dashboard","Happy","P1",A,
  "Signed out","—","1. Sign in as admin@kuja.org / pass123",
  "Lands on the operator/admin dashboard with platform tiles."),
 ("ADM-002","Users","Search and filter the user list","Happy","P2",A,
  "Signed in","—","1. Go to /admin/users\n2. Search a name/email; filter by role",
  "Filtered users list (GET /api/admin/users) with Name/Email/Role/Org/Last login/Status."),
 ("ADM-003","Users","Add a user (temp password shown once)","Happy","P1",A,
  "Signed in","—",
  "1. 'Add user'\n2. Enter full name, email, role, organisation\n3. Create",
  "User is created (POST /api/admin/users); a server-generated temp password is shown exactly once; user must change it on first sign-in."),
 ("ADM-004","Users","Reset a user's password","Happy","P2",A,
  "Signed in; a demo user (NOT a protected account)","—",
  "1. On a demo user row click 'Reset password' and confirm",
  "A new one-time password is shown once (POST /users/{id}/reset-password); the user is forced to set their own next login."),
 ("ADM-005","Users","Deactivate then reactivate (no hard delete)","Edge","P2",A,
  "Signed in; a demo user","—","1. 'Remove' a demo user\n2. Then 'Restore' them",
  "User is deactivated then reactivated (POST /deactivate, /reactivate); there is no hard delete (audit-chain preservation)."),
 ("ADM-006","Access control","Non-admin cannot reach /admin/users","Negative","P1",N,
  "Signed in as an NGO","—","1. As an NGO, navigate to /admin/users (and call GET /api/admin/users)",
  "Access is refused (403); no user list is returned."),
 ("ADM-007","2FA","Enrol TOTP with QR + recovery codes","Security","P1",A,
  "Signed in; on /admin/security","An authenticator app (Google/Microsoft Authenticator)",
  "1. 'Enrol now' (POST /totp/enroll/start) and scan the QR\n2. Enter the 6-digit code and 'Confirm'\n3. Save the recovery codes",
  "TOTP activates; 10 single-use recovery codes are shown once; status flips to enrolled."),
 ("ADM-008","2FA","Verify TOTP at login and disable it","Security","P2",A,
  "TOTP enrolled","The authenticator app","1. Sign out and back in; enter a TOTP code (or a recovery code)\n2. Later, disable TOTP with a current code",
  "Login accepts a valid TOTP or a (consumed) recovery code; disable requires a current code (POST /totp/disable)."),
 ("ADM-009","2FA","Register and use a passkey (WebAuthn)","Security","P2",A,
  "A platform/browser passkey authenticator","—",
  "1. Register a passkey (POST /webauthn/register/begin+finish)\n2. Use it to re-auth a sensitive action",
  "The passkey registers and a later authenticate returns a short-lived (5 min, single-use) re-auth token; the gate is a no-op for users with no passkey enrolled."),
 ("ADM-010","Audit chain","Verify integrity, browse and export","Security","P1",A,
  "Signed in; on /admin/audit-chain","—",
  "1. View the integrity badge (GET /verify)\n2. Browse recent entries\n3. 'Export chain' (export.jsonl)\n4. 'Re-verify'",
  "Integrity shows intact; recent rows paginate; the JSONL exports; re-verify re-walks the hash chain and confirms."),
 ("ADM-011","Access control","Audit chain is admin-only","Negative","P1",D,
  "Signed in as a donor","—","1. As a donor, open /admin/audit-chain (and call /api/audit-chain/verify)",
  "An access-denied card / 401-403 is shown; no audit data is exposed."),
 ("ADM-012","Cron & telemetry","Cron health shows freshness bands","Happy","P2",A,
  "Signed in","—","1. Go to /admin/cron-health",
  "Each registered cron shows a band: fresh / overdue / never (GET /api/cron/health)."),
 ("ADM-013","Cron & telemetry","AI telemetry by endpoint","Happy","P2",A,
  "Signed in","—","1. Go to /admin/ai-telemetry\n2. Switch the window (24h/3d/7d/30d)",
  "Per-endpoint failure rate, p50/p95, tokens and recent failures render (GET /api/admin/ai-telemetry)."),
 ("ADM-014","Cron & telemetry","AI quality + cost + forecast","Happy","P3",A,
  "Signed in","—","1. Open /admin/ai-quality, /admin/ai-cost and /admin/cost-ceiling",
  "Quality/false-confidence rollups and cost-by-tenant/user + spend forecast render; a cost ceiling can be set."),
 ("ADM-015","Cron & telemetry","Tenant + system health + integrity","Happy","P3",A,
  "Signed in","—","1. Open /admin/tenant-health and the integrity/system-health surfaces",
  "Health dashboards and an integrity check render without errors."),
 ("ADM-016","Webhooks","Register a webhook (secret shown once)","Happy","P2",A,
  "Signed in; a request-bin/test URL","—",
  "1. Go to /settings/webhooks\n2. 'Register' with URL + events\n3. Copy the signing secret",
  "Webhook is created (POST /api/webhooks); the signing secret is shown exactly once."),
 ("ADM-017","Webhooks","Test delivery + view delivery log","Happy","P2",A,
  "A registered webhook","—",
  "1. Click 'Test'\n2. 'Show delivery history'",
  "A test event is delivered (POST /webhooks/{id}/test); the delivery log shows status/ms/tries (GET /webhooks/{id}/deliveries)."),
 ("ADM-018","Webhooks","Fires on a real event (grant published)","Edge","P3",A,
  "A webhook subscribed to grant.published","—","1. Have a donor publish a grant (DON-010)",
  "A 'grant.published' delivery appears in the webhook log with a 2xx (or retried) status."),
 ("ADM-019","2FA","Admins-without-2FA is visible / enforceable","Security","P3",A,
  "Signed in","—","1. View the users-without-2FA / TOTP-enrolment tile\n2. (If KUJA_ENFORCE_ADMIN_2FA=true) an un-enrolled admin write is blocked",
  "The tile reflects 2FA coverage; with enforcement on, an admin lacking TOTP is 401'd on write actions (nag-then-enforce)."),
 ("ADM-020","Access control","Cron trigger endpoints are scheduler-only","Security","P3",A,
  "Signed in","—","1. Note that POST /api/cron/* endpoints have no login gate (external scheduler)",
  "Confirm these are not linkable in the UI and are only invoked by the scheduler; flag if any is reachable without protection (documentation/hardening check)."),
]
ANY = "(any account)"
XCUT_ROWS = [
 # --- Auth & session -------------------------------------------------------
 ("XC-001","Auth & session","Session persists then expires (8h)","Security","P2",N,
  "Signed in","—","1. Sign in\n2. Reload after a short idle\n3. (If feasible) leave idle beyond 8h",
  "Session survives reloads (remember me); after PERMANENT_SESSION_LIFETIME (8h) the session expires and re-login is required."),
 ("XC-002","Auth & session","Logout ends the session","Happy","P2",N,
  "Signed in","—","1. Click logout",
  "Session ends (POST /api/auth/logout); protected routes redirect to /login."),
 ("XC-003","Auth & session","Password policy on change (min 10)","Security","P2",N,
  "On /change-password or /settings/security","—",
  "1. Try a new password < 10 chars\n2. Try one equal to the current\n3. Try a valid new one",
  "Short or unchanged passwords are rejected; a valid >=10-char password is accepted."),
 ("XC-004","Auth & session","Login IP rate limit","Security","P3",ANY,
  "Signed out","—","1. Fire > 100 login attempts from one IP within 5 minutes",
  "Requests beyond the per-IP limit return 429 (throttled)."),
 ("XC-005","Auth & session","Email-enumeration guard","Security","P3",ANY,
  "Signed out","—","1. Attempt logins across > 15 distinct emails within 5 minutes",
  "The enumeration guard returns 429 after 15 distinct emails/5 min."),
 ("XC-006","Auth & session","Deactivated account is refused","Security","P2",A,
  "Admin has deactivated a demo user","—","1. Try to sign in as the deactivated demo user",
  "Sign-in is refused with 403 'Account is deactivated'."),
 ("XC-007","Auth & session","Forced-change gate blocks the API","Security","P1",N,
  "A user flagged must_change_password","—",
  "1. Sign in\n2. Without changing the password, try to open /grants or call any /api/*",
  "Every /api call (except the allowlisted change-password/logout/me) is 403'd until the password is rotated."),
 ("XC-008","Auth & session","Server enforces role even if UI leaks","Security","P1",N,
  "Signed in as an NGO","—",
  "1. As an NGO, call a donor-only endpoint directly (e.g. POST /api/grants/ create, or PATCH an application status)",
  "The server refuses with 403 (role_required fails closed) — the client-side 'role != ngo' check never grants real access."),
 ("XC-009","Auth & session","CSRF header required on JSON writes","Security","P2",N,
  "Signed in","—","1. Send a JSON POST/PUT without the X-Requested-With/CSRF header",
  "The write is refused; multipart uploads are the intended exception (they still enforce type/size)."),
 # --- File uploads (validation matrix) ------------------------------------
 ("XC-010","File uploads","Oversized upload is rejected (413)","Negative","P1",N,
  "On /apply Documents (or any upload)","99_edge_cases/oversized_20MB.pdf",
  "1. Upload the 20 MB file",
  "Rejected with 413 and a message like 'File too large … Maximum size is 16 MB.'; nothing stored."),
 ("XC-011","File uploads","Disallowed type: SVG rejected","Negative","P1",N,
  "On an upload field","99_edge_cases/image_with_script.svg",
  "1. Try to upload the .svg",
  "Rejected with 400 'File type not allowed' (svg is not in the allow-list); the embedded script never runs."),
 ("XC-012","File uploads","Disallowed type: HTML rejected","Negative","P2",N,
  "On an upload field","99_edge_cases/webpage_upload.html",
  "1. Try to upload the .html",
  "Rejected with 400 'File type not allowed'."),
 ("XC-013","File uploads","Empty file rejected","Negative","P2",N,
  "On an upload field","99_edge_cases/empty_file.pdf",
  "1. Try to upload the 0-byte PDF",
  "Rejected with 400 'File is empty or too small' (< 100 bytes)."),
 ("XC-014","File uploads","Renamed executable caught by magic bytes","Security","P1",N,
  "On an upload field","99_edge_cases/disguised_program.exe (rename to .pdf) or report.pdf.exe",
  "1. Rename disguised_program.exe to disguised.pdf and upload it\n2. Also try report.pdf.exe as-is",
  "The renamed file fails the magic-byte check → 400 'content does not match .pdf'; the .exe extension is rejected outright."),
 ("XC-015","File uploads","Text renamed .pdf is rejected","Negative","P2",N,
  "On an upload field","99_edge_cases/not_really_a_pdf.pdf",
  "1. Upload the text file that has a .pdf extension",
  "Rejected — header isn't %PDF (magic-byte mismatch) → 400."),
 ("XC-016","File uploads","Corrupt/unreadable PDF rejected","Negative","P2",N,
  "On an upload field","99_edge_cases/corrupt_truncated.pdf",
  "1. Upload the truncated PDF",
  "Rejected — no extractable text (PDF must yield >= 20 chars) → 400."),
 ("XC-017","File uploads","Allowed .txt is accepted (field may still restrict)","Edge","P3",N,
  "On an upload field","99_edge_cases/proposal_wrong_format.txt",
  "1. Upload the .txt where a document is expected",
  "Global rules ACCEPT .txt (in the allow-list, > 20 chars). NOTE: some fields advertise only PDF/DOCX in their picker; confirm the field-level expectation and log if inconsistent."),
 ("XC-018","File uploads","Odd filenames are sanitised","Edge","P3",N,
  "On an upload field","99_edge_cases/very_long_filename_*.pdf and تقرير_ميداني_عربي.docx",
  "1. Upload a file with a very long name\n2. Upload a file with an Arabic (RTL) name",
  "Both store safely (secure_filename + UUID); the app serves them without path/JS injection and displays the name correctly."),
 ("XC-019","File uploads","Upload rate limit","Security","P3",N,
  "On an upload field","Any small valid PDF",
  "1. Upload > 10 files within a minute",
  "Beyond ~10/min the endpoint throttles (429)."),
 # --- i18n & RTL -----------------------------------------------------------
 ("XC-020","i18n & RTL","All six languages switch","i18n","P2",N,
  "Signed in","—","1. Switch language to each of English, Arabic, French, Spanish, Kiswahili, Soomaali",
  "UI strings change for each; the choice persists (PUT /api/auth/language)."),
 ("XC-021","i18n & RTL","Arabic renders right-to-left","i18n","P1",N,
  "Signed in","—","1. Switch to Arabic and open /dashboard, /grants, /apply",
  "Layout flips to RTL (dir='rtl'); the /login and /change-password pages intentionally stay LTR."),
 ("XC-022","i18n & RTL","Untranslated strings fall back to English","i18n","P2",N,
  "Signed in","—","1. Switch to Spanish or Somali (which have translation gaps)\n2. Browse deep pages",
  "Missing keys show the English text, never a raw key like 'grant.browse_title'."),
 ("XC-023","i18n & RTL","Server-generated labels are localised","i18n","P2",N,
  "Signed in in Arabic","—","1. Trigger a server-generated title/notification (e.g. a status change)\n2. Read it in Arabic",
  "Server strings render via title_key + params with an English fallback — not a raw key; verify in-language in the browser."),
 # --- PWA / offline / mobile ----------------------------------------------
 ("XC-024","PWA & offline","Install banner + install","Happy","P3",N,
  "A supported browser (Chrome/Edge), not already installed","—",
  "1. Sign in and wait for 'Install Kuja for faster access'\n2. Click 'Install' (or 'Not now')",
  "The banner appears once; Install adds the PWA; 'Not now' hides it and the dismissal persists (localStorage)."),
 ("XC-025","PWA & offline","Offline navigation uses the cached shell","Edge","P2",N,
  "PWA installed / visited once","—","1. Load the app online\n2. Go offline\n3. Navigate to a cached page",
  "The service worker serves a cached shell (network-first navigations fall back to cache); an offline banner shows."),
 ("XC-026","PWA & offline","Offline edits queue and sync","Edge","P1",N,
  "Signed in; in a draft (application/report)","—",
  "1. Go offline\n2. Make an edit that POSTs/PUTs\n3. Reconnect",
  "The write queues in the IndexedDB outbox and drains oldest-first on reconnect; a 2xx clears it, a 4xx surfaces for review; no data lost."),
 ("XC-027","Mobile","No horizontal overflow at 360×800","Edge","P1",N,
  "Signed in; device/emulated 360×800","—","1. Open /dashboard, /grants, /apply, /reports at 360×800",
  "No horizontal scrollbar; content fits; the off-canvas copilot rail is clipped when closed (<= 380px)."),
 ("XC-028","Mobile","Sidebar/padding correct at 390","Edge","P2",N,
  "Signed in; emulated 390px","—","1. Open the nav/sidebar at 390px",
  "Sidebar content isn't clipped; document width doesn't exceed the viewport."),
 ("XC-029","Mobile","Slow-3G login doesn't leak credentials","Security","P2",N,
  "Throttled to slow 3G; on /login","—","1. Type email/password and press Enter BEFORE the page hydrates",
  "Submit is gated behind hydration; the password never appears in the URL/query string."),
 # --- Notifications --------------------------------------------------------
 ("XC-030","Notifications","In-app inbox: list, unread, mark read","Happy","P2",N,
  "Signed in; has notifications","—",
  "1. Open the notifications inbox\n2. Note the unread count\n3. Mark one read, then 'mark all'",
  "Unread-first list; unread count decrements; mark-read and mark-all work (GET/PUT /api/notifications/*)."),
 ("XC-031","Notifications","Preferences: channels per category","Edge","P3",N,
  "Signed in","—","1. Open /settings/notifications\n2. Review channels per category",
  "Categories (deadlines/reviews/compliance/decisions/…) map to channels; compliance defaults to in-app only; changes save."),
 ("XC-032","Notifications","Email fires on a decision","Edge","P2",D,
  "SendGrid configured; NGO has email channel on for decisions","—",
  "1. As donor, award/decline an application (DON-017)",
  "The NGO receives an in-app notification and (if the email channel is on and SendGrid configured) a decision email."),
 ("XC-033","Notifications","Digest cadence setting","Edge","P3",N,
  "Signed in","—","1. Set digest cadence to weekly then off",
  "The weekly summary digest respects the cadence (weekly/off)."),
 # --- Accessibility --------------------------------------------------------
 ("XC-034","Accessibility","Keyboard-only through the apply wizard","A11y","P2",N,
  "Signed in; on /apply","—","1. Complete a step using only Tab/Shift-Tab/Enter/Space",
  "All controls are reachable and operable by keyboard; focus is always visible."),
 ("XC-035","Accessibility","Form fields have labels; errors announced","A11y","P2",N,
  "On a form (login, apply, profile)","—","1. Inspect fields with a screen reader / accessibility tree",
  "Inputs have associated labels; validation errors are programmatically announced."),
 ("XC-036","Accessibility","Contrast holds in light and dark","A11y","P3",N,
  "Signed in","—","1. Toggle dark mode\n2. Check text/controls contrast on key pages",
  "Text and interactive elements meet contrast in both themes; no invisible-on-load text."),
 ("XC-037","Accessibility","Reduced motion respected","A11y","P3",N,
  "OS 'reduce motion' enabled","—","1. Browse animated surfaces",
  "Non-essential animation is reduced/disabled when prefers-reduced-motion is set."),
 # --- Performance ----------------------------------------------------------
 ("XC-038","Performance","Application submit stays within budget","Perf","P2",N,
  "Signed in; a complete application","—","1. Submit and time the response",
  "Submit returns within the latency budget; heavy AI work runs off the critical path (async), not blocking the submit."),
 ("XC-039","Performance","Dashboard tiles load without hanging","Perf","P2",ANY,
  "Signed in","—","1. Open each role's dashboard and drill into a tile",
  "Tiles resolve promptly; drill-ins don't hang or spin indefinitely."),
]

def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    readme(wb); accounts(wb); coverage(wb)
    add_case_sheet(wb, "NGO Journey", "2A9D8F",
        "Applicant end-to-end: login → org profile → browse → apply → decision → compliance/reporting", NGO_ROWS)
    add_case_sheet(wb, "Capacity & Trust", "8E7CC3",
        "In-app assessment + Trust Passport, standalone Kuja Trust app, and the Grant↔Trust seam", TRUST_ROWS)
    add_case_sheet(wb, "Donor Journey", "C1710C".replace("C1710C","B8620A"),
        "Grantmaker end-to-end: create/publish grant → review applications → decide → portfolio", DONOR_ROWS)
    add_case_sheet(wb, "Reviewer Journey", "3D5A80",
        "Independent reviewer: queue → score against rubric → COI/decline/snooze → AI tools", REVIEWER_ROWS)
    add_case_sheet(wb, "Admin & Platform", "5B5B5B",
        "Operator: users/provisioning → 2FA → audit chain → cron/AI telemetry → webhooks", ADMIN_ROWS)
    add_case_sheet(wb, "Cross-cutting", "7A3B69",
        "Auth/security, file-upload validation, i18n/RTL, PWA/offline/mobile, notifications, accessibility", XCUT_ROWS)
    defect_log(wb)
    wb.save(OUT)
    n = sum(len(x) for x in (NGO_ROWS,TRUST_ROWS,DONOR_ROWS,REVIEWER_ROWS,ADMIN_ROWS,XCUT_ROWS))
    print(f"Wrote {OUT}  ({n} test cases)")
    try:
        os.makedirs(DL, exist_ok=True); shutil.copy(OUT, os.path.join(DL, os.path.basename(OUT)))
        print("Copied to", DL)
    except Exception as e:
        print("copy skipped:", e)

def readme(wb):
    b = [
      ("title","Kuja Marketplace — User Acceptance Test (UAT) Plan"),
      ("para","Version 1.0 · 15 Aug 2026 · Scope: the KUJA MARKETPLACE tenant only "
              "(default network, slug='kuja'). Proximate / Saxansaxo / NEAR tenants are out of scope."),
      ("space",1),
      ("header","1 · What this pack is"),
      ("para","A ready-to-run acceptance test plan AND a team-onboarding tool. Each persona tab is a "
              "list of test cases with exact steps (log in as X, click Y, upload Z) and the expected "
              "result. Work top-to-bottom; a tester records Actual result / Status / Tester / Date / "
              "Notes for every row. New team members can read the tabs to learn how the product works."),
      ("bullet","Pack contents: this workbook + docs/uat/testfiles/ (48 realistic upload files) + "
                "docs/uat/UAT_Guide.md (facilitation notes)."),
      ("space",1),
      ("header","2 · Environments & URLs"),
      ("table",(["What","Where","Notes"],[
        ["Kuja Grant app (prod)","https://web-production-6f8a.up.railway.app","Primary system under test"],
        ["Kuja Grant app (branded)","(go-live domain — TBC)","No demo login buttons on a branded domain"],
        ["Kuja Trust app (live UAT)","https://kuja-app-production.up.railway.app","Capacity assessment / Passport"],
        ["Demo login buttons","localhost / *.up.railway.app / NEXT_PUBLIC_DEMO_MODE=true","Hidden on branded prod"],
      ])),
      ("para","Prefer a STAGING/UAT deployment for destructive cases. If testing on prod, do NOT create "
              "spam grants/applications under real accounts — use the demo accounts, and note anything "
              "that needs cleanup in the Defect Log."),
      ("space",1),
      ("header","3 · How to read a test case"),
      ("table",(["Column","Meaning"],[
        ["Test ID","Stable id, e.g. NGO-012. Reference it in the Defect Log."],
        ["Type","Happy = golden path · Edge = boundary · Negative = must be rejected · Security · Setup · Perf · i18n · A11y"],
        ["Pri","P1 = blocker for go-live · P2 = important · P3 = nice-to-have"],
        ["Login as","Which seeded account to use (see the Test Accounts tab)"],
        ["Test data / files","Which file(s) from docs/uat/testfiles/ to use"],
        ["Steps / Expected result","Do exactly this; you should see exactly that"],
        ["Actual / Status / Tester / Date / Notes","Filled in by the tester. Status is a dropdown."],
      ])),
      ("space",1),
      ("header","4 · Status legend"),
      ("bullet","Pass — behaves as the Expected result says."),
      ("bullet","Fail — differs from Expected. Log it in the Defect Log with the Test ID + a screenshot."),
      ("bullet","Blocked — could not run (dependency/env). Say why in Notes."),
      ("bullet","Not run / In progress — self-explanatory."),
      ("space",1),
      ("header","5 · Ground rules"),
      ("warn","NEVER sign in as, reset, or modify the 6 REAL go-live accounts (iloyan@, mrashid@, "
              "thussein@adesoafrica.org; mtumwebaze@adesoafrica.org; kali@, msattar@proximatefund.org). "
              "Use ONLY the demo accounts on the Test Accounts tab (all password: pass123)."),
      ("warn","Never enter a real password, bank/card number, or government ID into any field. "
              "All test files are fictional; the org 'Amani Health Initiative' and donor 'Global Health "
              "Fund' do not exist."),
      ("bullet","Two people are ideal for review-cycle cases (one NGO + one donor/reviewer) so an "
                "application can move submitted → reviewed → decided in one sitting."),
      ("bullet","Clear browser storage between mobile/PWA/i18n cases (they persist choices in localStorage)."),
    ]
    add_freeform(wb, "READ ME", "1F6F3C", b)

def accounts(wb):
    b = [
      ("title","Test accounts & data"),
      ("para","All demo accounts use password: pass123. On demo hosts the login page shows one-click "
              "buttons for the NGO/Donor/Reviewer trio; otherwise type the email + pass123."),
      ("space",1),
      ("header","Seeded demo accounts (Kuja Marketplace)"),
      ("table",(["Email","Password","Role","Organisation","Use for"],[
        ["fatima@amani.org","pass123","ngo","Amani Foundation","Primary NGO applicant journey"],
        ["ahmed@salamrelief.org","pass123","ngo","Salam Relief","2nd NGO (compare / peer data)"],
        ["thandi@ubuntu.org","pass123","ngo","Ubuntu","3rd NGO"],
        ["peter@hopebridges.org","pass123","ngo","Hope Bridges","4th NGO"],
        ["aisha@sahelwomen.org","pass123","ngo","Sahel Women","5th NGO"],
        ["sarah@globalhealth.org","pass123","donor","Global Health Fund","Primary donor journey"],
        ["david@eatrust.org","pass123","donor","East Africa Dev. Trust","2nd donor"],
        ["james@reviewer.org","pass123","reviewer","Independent Review Assoc.","Primary reviewer journey"],
        ["maria@reviewer.org","pass123","reviewer","Independent Review Assoc.","2nd reviewer (panel/COI)"],
        ["admin@kuja.org","pass123","admin","(platform)","Admin/operator journey"],
      ])),
      ("space",1),
      ("header","Kuja Trust app (standalone) sign-in"),
      ("bullet","Easiest: on https://kuja-app-production.up.railway.app click 'Explore the demo "
                "workspace' — one click, no credentials (binds to a seeded demo org)."),
      ("bullet","Named Trust logins are provisioned out-of-band (governance key); certification personas "
                "must use @kuja-test.invalid emails. There is no shipped username/password for Trust."),
      ("space",1),
      ("header","Protected — do not touch"),
      ("warn","iloyan@adesoafrica.org · mrashid@adesoafrica.org · thussein@adesoafrica.org · "
              "mtumwebaze@adesoafrica.org · kali@proximatefund.org · msattar@proximatefund.org — "
              "these are REAL go-live accounts. Never log in as, reset, or alter them."),
      ("space",1),
      ("header","Test-file pack (docs/uat/testfiles/)"),
      ("table",(["Folder","Contains"],[
        ["01_registration","Reg. certificate, constitution, audited financials, board list, tax exemption, org chart"],
        ["02_capacity_trust_evidence","Safeguarding/PSEA, financial, anti-fraud, HR, data-protection policies; bank & donor-reference letters"],
        ["03_grant_application","Proposal (PDF+DOCX), budget (XLSX+PDF), logframe, workplan, MOU, letter of support"],
        ["04_compliance_reporting","Narrative report (DOCX), financial report (XLSX), receipts, field photos (JPG/WebP), attendance, video + voice memo"],
        ["05_donor","Call-for-Proposals PDF (for the AI-extraction wizard) + reviewer scoring rubric"],
        ["99_edge_cases","oversized_20MB.pdf, empty, corrupt, renamed .exe, double-extension, SVG/HTML, active-JavaScript PDF, password-protected PDF, Arabic filename"],
      ])),
    ]
    add_freeform(wb, "Test Accounts", "0B7285", b)

def coverage(wb):
    sets = [("NGO Journey",NGO_ROWS),("Capacity & Trust",TRUST_ROWS),("Donor Journey",DONOR_ROWS),
            ("Reviewer Journey",REVIEWER_ROWS),("Admin & Platform",ADMIN_ROWS),("Cross-cutting",XCUT_ROWS)]
    types = ["Happy","Edge","Negative","Security","Setup","Perf","i18n","A11y"]
    data = []
    for name, rows in sets:
        counts = {t:0 for t in types}
        for r in rows: counts[r[3]] = counts.get(r[3],0)+1
        data.append([name, str(len(rows))] + [str(counts.get(t,0) or "") for t in types])
    total = ["TOTAL", str(sum(len(r) for _,r in sets))] + \
            [str(sum(1 for _,rows in sets for x in rows if x[3]==t) or "") for t in types]
    b = [
      ("title","Coverage summary"),
      ("para","Test-case counts per persona tab and per type. Use it to plan a UAT session and to show "
              "the team the breadth of coverage. P1 blockers should be run first in every tab."),
      ("space",1),
      ("table",(["Persona / area","Total"]+types, data + [total])),
      ("space",1),
      ("header","Journeys covered end-to-end"),
      ("bullet","NGO: sign in → complete org profile → browse/search/watchlist → apply (eligibility → "
                "proposal + AI aids → documents → submit gate) → track decision → withdraw/appeal/revise → "
                "post-award compliance & reporting (voice, photo evidence, pre-check, submit)."),
      ("bullet","Capacity & Trust: in-app assessment + Passport publish/share/verify; the standalone Kuja "
                "Trust app (7 C4C domains, evidence room, screening, self-declared vs verified-clear, "
                "languages/RTL, low-data mode); and the LIVE Grant↔Trust hand-off + service read-back."),
      ("bullet","Donor: create grant (incl. PDF→AI extraction) → set criteria/weights → publish (licence) "
                "→ review pipeline → request revision/documents → shortlist/compare → decide → debrief → "
                "appeals → portfolio Q&A → review grantee reports → exports."),
      ("bullet","Reviewer: queue → score against rubric → AI triage/score/rationale → COI auto-recuse → "
                "decline/snooze. Admin: users/provisioning → 2FA → audit chain → cron/AI telemetry → webhooks."),
      ("bullet","Cross-cutting: auth lockout/session/forced-password-change; the full file-upload "
                "validation matrix; i18n across en/ar/fr/es/sw/so with Arabic RTL; PWA/offline/mobile; "
                "notifications/email; accessibility."),
    ]
    add_freeform(wb, "Coverage", "334155", b)

def defect_log(wb):
    ws = wb.create_sheet("Defect Log"); ws.sheet_properties.tabColor = "8A2A2A"
    hdr = ["Defect ID","Test ID","Date","Severity","Persona","Summary","Steps to reproduce",
           "Expected","Actual","Environment","Screenshot ref","Reported by","Status","Owner","Fix notes"]
    w   = [11,10,11,10,12,28,40,28,28,16,16,14,12,14,30]
    _title(ws, "Defect Log — raise one row per failed test case (reference the Test ID)", len(hdr))
    for j,h in enumerate(hdr, start=1):
        c = ws.cell(2,j,h); c.font=Font(bold=True,color=WHITE,size=10)
        c.fill=PatternFill("solid",fgColor=NAVY2); c.border=BORDER
        c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
    for j,wd in enumerate(w, start=1): ws.column_dimensions[get_column_letter(j)].width = wd
    ws.row_dimensions[2].height = 24
    # a couple of example rows (greyed) then blanks
    ex = ["D-001","NGO-031","","P1 / Blocker","NGO",
          "Report attachment Upload button 404s","Open /reports, expand a draft report, click Upload, choose a PDF",
          "File attaches to the report","POST /api/reports/{id}/attachments returns 404 (route not implemented)",
          "prod","(attach)","<name>","Open","",""]
    for j,v in enumerate(ex, start=1):
        c=ws.cell(3,j,v); c.border=BORDER; c.font=Font(size=9,italic=True,color=GREY)
        c.alignment=Alignment(vertical="top",wrap_text=True)
    for r in range(4, 60):
        for j in range(1, len(hdr)+1):
            c=ws.cell(r,j,""); c.border=BORDER; c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{get_column_letter(len(hdr))}59"
    dv=DataValidation(type="list",formula1='"Open,In progress,Fixed,Retest,Closed,Won\'t fix"',allow_blank=True)
    ws.add_data_validation(dv); dv.add("M3:M59")

if __name__ == "__main__":
    build()
