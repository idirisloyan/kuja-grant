#!/usr/bin/env python3
"""
generate_review_workbooks.py — one native-speaker review workbook per language.

Reads the app's i18n string tables (frontend/src/i18n/<lang>.json — flat dotted
keys) and produces a reviewer-friendly .xlsx per language containing every UI
string, the English source, the current translation, and editable columns for
the reviewer's corrections. English drives the key list (it is the source of
truth); a key missing from a translation is flagged "MISSING", and a translation
identical to English is flagged "SAME AS ENGLISH — check".

Outputs to BOTH the repo (docs/i18n-review/) and the shared Downloads folder.

Run from the repo root:  py -3 docs/i18n-review/generate_review_workbooks.py
"""
import json
import os
import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, '..', '..'))
I18N = os.path.join(REPO, 'frontend', 'src', 'i18n')
DOWNLOADS = r'C:\Users\IdirisLoyan\Downloads\Latest Kuja files'
STAMP = date.today().isoformat()  # informational only

# This review is scoped to the KUJA MARKETPLACE tenant only. The i18n files are
# one shared table serving all four tenants, so we drop the namespaces that only
# appear in the other tenants' products (Proximate's endorsement/disbursement/
# crisis vocabulary — all under `proximate.*` + `prox_grant`; NEAR's networked-
# funds vocabulary — declarations, network membership, window/near reports).
# Everything else (grants, applications, donors, reviewers, NGOs, Trust Profile,
# compliance, reporting, AI, dashboards, admin) is shared and reviewed.
EXCLUDE_NAMESPACES = {
    'proximate', 'prox_grant',
    'declaration', 'declaration_detail',
    'membership_detail', 'membership_list', 'network_join',
    'window_report', 'near_reports',
    'crisis_detail', 'crisis_list',
}

# (code, English name, native name, is_RTL)
LANGS = [
    ('en', 'English', 'English', False),
    ('so', 'Somali', 'Soomaali', False),
    ('sw', 'Swahili', 'Kiswahili', False),
    ('es', 'Spanish', 'Espanol', False),
    ('fr', 'French', 'Francais', False),
    ('ar', 'Arabic', 'Arabic', True),
]

# Placeholders / markup the reviewer must copy through UNCHANGED.
PLACEHOLDER_RE = re.compile(r'(\{[^}]*\}|%\((?:\w+)\)s|%[sd]|\$\{[^}]*\}|<[^>]+>|&[a-z]+;)')

# App-domain glossary — reviewers fill the "Agreed term" column and reuse it
# everywhere for consistency. English + short definition only (we do not guess
# translations); a few Arabic terms are pre-agreed with the client.
GLOSSARY = [
    ('Grant', 'A funding opportunity a donor publishes for NGOs to apply to.'),
    ('Funding window / round', 'A time-boxed call within a fund.'),
    ('Application', "An NGO's submission against a grant."),
    ('Applicant / NGO', 'The organisation applying for funding.'),
    ('Donor', 'The organisation offering the grant.'),
    ('Reviewer', 'Independent scorer of applications.'),
    ('Oversight Body (OB)', 'Governance body that approves/gates decisions.'),
    ('Trust Profile / Capacity Passport', "An org's reusable due-diligence profile."),
    ('Due diligence', 'Vetting checks (registration, audit, sanctions).'),
    ('Sanctions screening', 'Check against sanctions/exclusion lists.'),
    ('Compliance', 'Meeting the required rules/documents.'),
    ('Criteria / Rubric', 'The weighted scoring dimensions of a grant.'),
    ('Score', 'A rating against the criteria.'),
    ('Award / Awarded', 'The decision to fund an application.'),
    ('Disbursement', 'Payment of grant funds to a partner.'),
    ('Deadline', 'The date by which something is due.'),
    ('Submit / Submission', 'Send an application/report for review.'),
    ('Report / Reporting', 'Post-award progress and financial reporting.'),
    ('Appeal', 'A request to reconsider a decision.'),
    ('Dashboard', "The user's home overview screen."),
    ('Draft', 'Unsubmitted, private work-in-progress.'),
]
# Pre-agreed Arabic terms (from the Proximate client sign-off — keep consistent).
GLOSSARY_AR_PREFILL = {
    'Oversight Body (OB)': 'هيئة الإشراف',
    'Funding window / round': 'دورة',
    'Appeal': 'تظلّم / استئناف',
}

# ---- styles -------------------------------------------------------------
HDR_FILL = PatternFill('solid', fgColor='1F4E5F')
HDR_FONT = Font(color='FFFFFF', bold=True, size=11)
REF_FILL = PatternFill('solid', fgColor='EEF2F4')     # read-only reference cols
EDIT_FILL = PatternFill('solid', fgColor='FFF7D6')    # reviewer edits here
MISS_FILL = PatternFill('solid', fgColor='FDE2E1')    # missing translation
SAME_FILL = PatternFill('solid', fgColor='FCEFCB')    # same-as-english
TITLE_FONT = Font(bold=True, size=16, color='1F4E5F')
H2_FONT = Font(bold=True, size=12, color='1F4E5F')
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='D0D7DA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(code):
    with open(os.path.join(I18N, f'{code}.json'), encoding='utf-8') as fh:
        return json.load(fh)


def placeholders(s):
    return '  '.join(dict.fromkeys(PLACEHOLDER_RE.findall(s or '')))


def build(code, name, native, rtl, en):
    is_src = (code == 'en')
    tgt = en if is_src else load(code)
    keys = sorted(k for k in (set(en) | set(tgt))
                  if k.split('.')[0] not in EXCLUDE_NAMESPACES)

    wb = Workbook()

    # ---------- Instructions tab ----------
    ins = wb.active
    ins.title = 'READ ME FIRST'
    ins.sheet_view.showGridLines = False
    ins.column_dimensions['A'].width = 3
    ins.column_dimensions['B'].width = 118
    def line(row, text, font=None, fill=None):
        c = ins.cell(row=row, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if font: c.font = font
        if fill: c.fill = fill
        return row + 1
    r = 2
    r = line(r, f'Kuja — {name} ({native}) translation review', TITLE_FONT)
    r += 1
    if is_src:
        r = line(r, 'You are reviewing the ENGLISH source copy. English is the master text every other '
                    'language is translated from, so please proofread it for clarity, correctness, tone and '
                    'consistency. Improvements you make here should also be reflected in the other languages.', )
    else:
        r = line(r, f'Thank you for reviewing the {name} translation of the Kuja Marketplace. These are the exact '
                    f'words end-users (NGOs, donors, reviewers, admins) see in the app. Your job is to make the '
                    f'{name} read naturally, correctly and respectfully for a grassroots / humanitarian audience. '
                    f'(Scope: this file covers the Kuja Marketplace product only — strings that belong to the other '
                    f'Kuja networks have been left out so you do not have to review them.)')
    r += 1
    r = line(r, 'What we need you to do', H2_FONT)
    r = line(r, '1.  Go to the "Strings to review" tab.')
    if is_src:
        r = line(r, '2.  Read each row. If the English can be clearer or is wrong, type the better wording in the '
                    'yellow "Your corrected text" column. If it is already good, leave it blank and put "Y" in Approved.')
    else:
        r = line(r, '2.  For each row compare the English (source) with the current translation. If the translation '
                    'is wrong, unnatural, or missing, type your corrected translation in the yellow '
                    '"Your corrected translation" column. If it is already good, leave it blank and put "Y" in Approved.')
    r = line(r, '3.  Add anything you want us to know in "Reviewer notes" (e.g. "no space in UI", "ambiguous").')
    r = line(r, '4.  Use the "Glossary" tab to keep key terms translated the SAME way everywhere.')
    r += 1
    r = line(r, 'Critical rules (please do not skip)', H2_FONT)
    r = line(r, '•  KEEP PLACEHOLDERS EXACTLY. Text in { } such as {name}, {count}, {days}; %s / %d; and HTML tags '
                'like <b>…</b> are filled in by the software. Copy them through unchanged and only move them if the '
                'grammar of the sentence requires it. The "Placeholders to keep" column lists the ones in each row.', EDIT_FILL)
    r = line(r, '•  Keep it SHORT. These are buttons, labels and messages with limited space. Prefer the concise, '
                'natural phrasing over a literal word-for-word translation.')
    r = line(r, '•  Tone: clear, plain, respectful. Users may be first-time, low-bandwidth, non-expert. Avoid jargon.')
    r = line(r, '•  Do NOT translate: brand/product names (Kuja, Adeso, Proximate), and leave numbers, dates and '
                'currency symbols as they are.')
    if rtl:
        r = line(r, '•  Arabic is RIGHT-TO-LEFT. This sheet is set RTL. Keep Latin names/URLs/placeholders LTR within '
                    'the sentence. Use the client-agreed terms shown in the Glossary tab (e.g. هيئة الإشراف, دورة).')
    r += 1
    r = line(r, 'How rows are flagged (colour + "Status" column)', H2_FONT)
    r = line(r, '   MISSING (pink)  = no translation exists yet — please provide one.', None, MISS_FILL)
    if not is_src:
        r = line(r, '   SAME AS ENGLISH (amber) = the current value is identical to English — usually means it was '
                    'never translated; please check.', None, SAME_FILL)
    r += 1
    r = line(r, 'Tips', H2_FONT)
    r = line(r, '•  The "Key (context id)" column is the internal name of the string. If you are unsure what a string '
                'means, you can search that key in the live app, or ask us for a screenshot.')
    r = line(r, '•  You can filter (Data ▸ Filter is already on) by the "Status" column to see only MISSING rows first.')
    r += 1
    r = line(r, 'Returning your review', H2_FONT)
    r = line(r, 'Save this file (keep the file name) and send it back to the Kuja team. Please also tell us the '
                'target date you can return it by. Contact: [add reviewer coordinator email].')
    r += 1
    r = line(r, f'Total strings in this file: {len(keys)}   |   Source of truth: frontend/src/i18n/{code}.json',
             Font(italic=True, size=9, color='777777'))

    # ---------- Strings tab ----------
    ws = wb.create_sheet('Strings to review')
    if rtl:
        ws.sheet_view.rightToLeft = True
    cur_hdr = 'English (source) — edit to improve' if is_src else f'Current {name} translation'
    corr_hdr = 'Your corrected text' if is_src else 'Your corrected translation'
    headers = ['#', 'Section', 'Key (context id)', 'English (source)']
    if not is_src:
        headers.append(cur_hdr)
    headers += ['Placeholders to keep', 'Status', corr_hdr, 'Reviewer notes', 'Approved? (Y/N)']

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(wrap_text=True, vertical='center',
                                horizontal='right' if rtl else 'left')
        c.border = BORDER
    ws.freeze_panes = 'D2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # column index helpers
    idx = {h: i + 1 for i, h in enumerate(headers)}
    edit_cols = {idx[corr_hdr], idx['Reviewer notes'], idx['Approved? (Y/N)']}

    n_missing = n_same = 0
    row = 2
    for k in keys:
        en_v = en.get(k, '')
        tg_v = tgt.get(k, '')
        if is_src:
            status = ''
        elif k not in tgt or (tg_v or '').strip() == '':
            status = 'MISSING'; n_missing += 1
        elif (tg_v or '').strip() == (en_v or '').strip():
            status = 'SAME AS ENGLISH'; n_same += 1
        else:
            status = 'translated'
        vals = {
            '#': row - 1,
            'Section': k.split('.')[0],
            'Key (context id)': k,
            'English (source)': en_v,
            cur_hdr: tg_v,
            'Placeholders to keep': placeholders(en_v),
            'Status': status,
            corr_hdr: '',
            'Reviewer notes': '',
            'Approved? (Y/N)': '',
        }
        for h in headers:
            c = ws.cell(row=row, column=idx[h], value=vals[h])
            c.alignment = WRAP_TOP if not rtl else Alignment(wrap_text=True, vertical='top',
                                                             horizontal='right' if h in (cur_hdr, corr_hdr) else 'general')
            c.border = BORDER
            if idx[h] in edit_cols:
                c.fill = EDIT_FILL
            else:
                c.fill = REF_FILL
        if status == 'MISSING':
            ws.cell(row=row, column=idx['Status']).fill = MISS_FILL
        elif status == 'SAME AS ENGLISH':
            ws.cell(row=row, column=idx['Status']).fill = SAME_FILL
        row += 1

    widths = {'#': 6, 'Section': 16, 'Key (context id)': 38, 'English (source)': 52,
              cur_hdr: 52, 'Placeholders to keep': 20, 'Status': 16,
              corr_hdr: 52, 'Reviewer notes': 30, 'Approved? (Y/N)': 12}
    for h in headers:
        ws.column_dimensions[get_column_letter(idx[h])].width = widths[h]

    dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'{get_column_letter(idx["Approved? (Y/N)"])}2:'
           f'{get_column_letter(idx["Approved? (Y/N)"])}{row - 1}')

    # ---------- Glossary tab ----------
    gl = wb.create_sheet('Glossary')
    if rtl:
        gl.sheet_view.rightToLeft = True
    gh = ['English term', 'Meaning / context', f'Agreed {name} term (fill in & reuse)']
    for ci, h in enumerate(gh, 1):
        c = gl.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BORDER
    for ri, (term, mean) in enumerate(GLOSSARY, start=2):
        pre = '' if is_src else (GLOSSARY_AR_PREFILL.get(term, '') if code == 'ar' else '')
        for ci, val in enumerate([term, mean, pre], 1):
            c = gl.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_TOP; c.border = BORDER
            c.fill = EDIT_FILL if ci == 3 else REF_FILL
    gl.column_dimensions['A'].width = 30
    gl.column_dimensions['B'].width = 60
    gl.column_dimensions['C'].width = 34
    gl.freeze_panes = 'A2'
    if code == 'ar':
        note = gl.cell(row=len(GLOSSARY) + 3, column=1,
                       value='Pre-filled Arabic terms are already agreed with the client — please keep them.')
        note.font = Font(italic=True, color='777777')

    # ---------- save to both locations ----------
    fname = f'Kuja_i18n_review_{code}_{name}.xlsx'
    out_repo = os.path.join(HERE, fname)
    out_dl = os.path.join(DOWNLOADS, fname)
    wb.save(out_repo)
    try:
        wb.save(out_dl)
    except Exception as e:
        print(f'  ! could not write to Downloads: {e}')
    return fname, len(keys), n_missing, n_same


def main():
    en = load('en')
    print(f'{"lang":10} {"strings":>8} {"missing":>8} {"same-as-en":>11}   file')
    for code, name, native, rtl in LANGS:
        fname, total, miss, same = build(code, name, native, rtl, en)
        print(f'{name:10} {total:8d} {miss:8d} {same:11d}   {fname}')
    print(f'\nWritten to:\n  repo:      {HERE}\n  downloads: {DOWNLOADS}')


if __name__ == '__main__':
    main()
