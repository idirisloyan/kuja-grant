"""
Generate test files for Kuja Grant v3.0 Test Cases
Creates positive and negative test files for comprehensive testing.
"""
import os
import sys
import struct
import zipfile
from datetime import datetime

# Paths
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DIR = os.path.join(BASE, 'test-files-v3')
DOCS_DIR = os.path.join(BASE, 'docs')
os.makedirs(TEST_DIR, exist_ok=True)

# ── Check for fpdf2 / openpyxl ──────────────────────────────────────
try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False
    print("Warning: fpdf2 not installed. Using minimal PDF generation.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Skipping XLSX generation.")


# ── Minimal PDF generator (fallback) ────────────────────────────────
def make_minimal_pdf(filepath, pages):
    """Create a valid PDF with text content using fpdf2 or raw bytes."""
    if HAS_FPDF:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(15, 15, 15)
        for page_content in pages:
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 16)
            title = page_content.get('title', '')
            pdf.cell(w=0, h=12, text=title, new_x='LMARGIN', new_y='NEXT', align='C')
            pdf.ln(5)
            pdf.set_font('Helvetica', '', 10)
            body = page_content.get('body', '')
            for line in body.split('\n'):
                line = line.strip()
                if not line:
                    pdf.ln(4)
                    continue
                if line.startswith('## '):
                    pdf.set_font('Helvetica', 'B', 13)
                    pdf.cell(w=0, h=8, text=line[3:], new_x='LMARGIN', new_y='NEXT')
                    pdf.set_font('Helvetica', '', 10)
                else:
                    # Truncate very long lines to avoid rendering issues
                    if len(line) > 500:
                        line = line[:500] + '...'
                    try:
                        pdf.multi_cell(w=0, h=5, text=line)
                    except Exception:
                        pdf.cell(w=0, h=5, text=line[:80], new_x='LMARGIN', new_y='NEXT')
        pdf.output(filepath)
    else:
        # Absolute minimal valid PDF
        content = '\n'.join(p.get('title', '') + '\n' + p.get('body', '') for p in pages)
        raw = (
            b'%PDF-1.4\n'
            b'1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
            b'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n'
            b'3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]'
            b'/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n'
            b'5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n'
        )
        stream = f'BT /F1 10 Tf 72 720 Td ({content[:500]}) Tj ET'.encode()
        stream_obj = (
            f'4 0 obj<</Length {len(stream)}>>\nstream\n'.encode()
            + stream
            + b'\nendstream\nendobj\n'
        )
        xref_offset = len(raw) + len(stream_obj)
        raw += stream_obj
        raw += (
            b'xref\n0 6\n'
            b'0000000000 65535 f \n'
            b'0000000009 00000 n \n'
            b'0000000058 00000 n \n'
            b'0000000115 00000 n \n'
            b'0000000300 00000 n \n'
            b'0000000250 00000 n \n'
            b'trailer<</Size 6/Root 1 0 R>>\n'
            + f'startxref\n{xref_offset}\n%%EOF\n'.encode()
        )
        with open(filepath, 'wb') as f:
            f.write(raw)


# ═══════════════════════════════════════════════════════════════════
# POSITIVE TEST FILES
# ═══════════════════════════════════════════════════════════════════

def create_valid_financial_report():
    """1. Valid financial report with proper structure."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_financial_report.pdf'), [
        {
            'title': 'ANNUAL FINANCIAL REPORT 2024-2025',
            'body': '''Amani Community Development Organization
Nairobi, Kenya | Registration: NGO/2012/3847

## Executive Summary
This financial report covers the fiscal year ending December 31, 2025. Total revenue for the period was $2,450,000 across four major grants and institutional funding. Total expenditure was $2,312,000, resulting in a surplus of $138,000.

## Revenue Summary
- Grant Income: $1,850,000 (75.5%)
- Institutional Funding: $400,000 (16.3%)
- Earned Revenue: $150,000 (6.1%)
- Other Income: $50,000 (2.0%)
Total Revenue: $2,450,000

## Expenditure by Category
- Personnel Costs: $920,000 (39.8%)
- Program Activities: $780,000 (33.7%)
- Travel and Transport: $185,000 (8.0%)
- Equipment and Supplies: $142,000 (6.1%)
- Office and Administrative: $165,000 (7.1%)
- Overhead and Indirect: $120,000 (5.2%)
Total Expenditure: $2,312,000

## Budget vs. Actual Analysis
Overall budget utilization rate: 94.2%
Variance: -5.8% (under-budget)
Key variance factors: delayed procurement cycle in Q3, favorable exchange rates.

## Internal Controls
- Segregation of duties maintained across finance team (4 staff)
- Monthly bank reconciliations completed on time
- Quarterly internal audits conducted
- Procurement thresholds enforced ($5,000 competitive bidding)
- Petty cash reconciled weekly

## Cash Flow Statement
Opening Balance (Jan 1, 2025): $342,000
Cash Inflows: $2,450,000
Cash Outflows: $2,312,000
Closing Balance (Dec 31, 2025): $480,000

## Donor Compliance
All donor-specific reporting requirements met. No outstanding questioned costs. Grant closeout procedures initiated for 2 completed grants.
'''
        },
        {
            'title': 'INDEPENDENT AUDITOR STATEMENT',
            'body': '''Baker Tilly Kenya - Certified Public Accountants
Audit Opinion: Unqualified (Clean)

We have audited the financial statements of Amani Community Development for the year ended December 31, 2025. In our opinion, the financial statements present fairly, in all material respects, the financial position of the organization and are in accordance with International Financial Reporting Standards (IFRS).

Basis for Opinion:
We conducted our audit in accordance with International Standards on Auditing (ISA). We are independent of the organization and have fulfilled our ethical responsibilities.

Key Audit Matters:
- Grant revenue recognition tested and confirmed
- Procurement compliance verified
- Fixed asset register reconciled
- Payroll tested for accuracy and authorization

Signed: John Mwangi, CPA
Date: February 28, 2026
License: ICPAK/2845
'''
        }
    ])
    print("  Created: valid_financial_report.pdf")


def create_valid_registration_cert():
    """2. Valid registration certificate."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_registration_cert.pdf'), [
        {
            'title': 'REPUBLIC OF KENYA - CERTIFICATE OF REGISTRATION',
            'body': '''NGO Coordination Board
Ministry of Interior and Coordination of National Government

CERTIFICATE OF REGISTRATION
Under the Non-Governmental Organizations Co-ordination Act, 1990

This is to certify that:

Organization Name: Amani Community Development
Registration Number: NGO/2012/3847
Date of Registration: March 15, 2012
Date of Expiry: March 15, 2027
Status: ACTIVE

Registered Office: Westlands Commercial Centre, 4th Floor
P.O. Box 12345-00100, Nairobi, Kenya

Objectives:
1. To promote community health and nutrition programs
2. To strengthen water, sanitation, and hygiene infrastructure
3. To build capacity of local health workers and community organizations

Board of Directors:
- Chairperson: Dr. Amina Osman
- Secretary: James Kariuki
- Treasurer: Grace Wanjiku
- Members: 4 additional directors

This certificate is valid subject to compliance with the provisions of the Act and regulations made thereunder.

Issued by: NGO Coordination Board
Authorized Signatory: _________________________
Date: March 15, 2022 (Renewal)
Official Stamp: [GOVERNMENT OF KENYA SEAL]
'''
        }
    ])
    print("  Created: valid_registration_cert.pdf")


def create_valid_audit_report():
    """3. Valid audit report."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_audit_report.pdf'), [
        {
            'title': 'INDEPENDENT AUDIT REPORT - FISCAL YEAR 2024',
            'body': '''Baker Tilly Kenya LLP
Certified Public Accountants
Nairobi, Kenya

To the Board of Directors
Amani Community Development

## Report on the Audit of Financial Statements

## Opinion
We have audited the financial statements of Amani Community Development, which comprise the statement of financial position as at December 31, 2024, and the statement of comprehensive income, statement of changes in funds, and statement of cash flows for the year then ended, and notes to the financial statements.

In our opinion, the accompanying financial statements present fairly, in all material respects, the financial position of the organization as at December 31, 2024, and its financial performance and cash flows for the year then ended in accordance with International Financial Reporting Standards (IFRS).

## Basis for Opinion
We conducted our audit in accordance with International Standards on Auditing (ISAs). Our responsibilities under those standards are further described below. We are independent of the organization in accordance with the International Ethics Standards Board for Accountants.

## Key Audit Matters
1. Revenue Recognition: Grant revenue of $1.85M properly recognized based on conditions met.
2. Fixed Assets: Physical verification conducted; register reconciled. Depreciation rates consistent.
3. Procurement: Sample testing of 45 transactions confirmed compliance with procurement policy.
4. Payroll: Tested 100% of payroll for 3 months. All correctly calculated and authorized.

## Internal Control Findings
Finding 1 (Low Risk): Petty cash reconciliation occasionally delayed by 2-3 days.
Recommendation: Implement daily reconciliation checklist.
Management Response: Accepted. Will implement by Q2 2025.

## Going Concern
No material uncertainties exist regarding the organization\'s ability to continue as a going concern.

Signed: John Mwangi, CPA
Partner, Baker Tilly Kenya LLP
ICPAK License: 2845
Date: March 15, 2025
'''
        }
    ])
    print("  Created: valid_audit_report.pdf")


def create_valid_psea_policy():
    """4. Valid PSEA policy document."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_psea_policy.pdf'), [
        {
            'title': 'PROTECTION FROM SEXUAL EXPLOITATION AND ABUSE (PSEA) POLICY',
            'body': '''Amani Community Development
Policy Version: 3.0 | Effective: January 1, 2025
Approved by: Board of Directors | Review Date: January 1, 2027

## 1. Policy Statement
Amani Community Development has zero tolerance for sexual exploitation and abuse (SEA). All personnel, including staff, volunteers, consultants, and partners, are required to adhere to this policy. Any violation will result in disciplinary action, up to and including termination and referral to law enforcement.

## 2. Definitions
- Sexual Exploitation: Any actual or attempted abuse of vulnerability, differential power, or trust for sexual purposes.
- Sexual Abuse: Actual or threatened physical intrusion of a sexual nature, whether by force or under unequal conditions.
- Beneficiary: Any person who receives services or assistance from the organization.
- Survivor: Person who has experienced SEA.

## 3. Standards of Conduct
All personnel shall:
- Not engage in sexual exploitation or abuse
- Not engage in sexual activity with persons under 18
- Not exchange money, employment, goods, or services for sex
- Report any concerns or suspicions of SEA
- Maintain professional boundaries with beneficiaries
- Complete mandatory PSEA training annually

## 4. Reporting Mechanism
- Confidential reporting channel: psea@amani.org
- Anonymous hotline: +254-XXX-XXXX (toll-free)
- Reports may also be made to any manager or HR staff
- External reporting to donor PSEA focal points
- All reports treated confidentially

## 5. Investigation Procedures
- All complaints investigated within 48 hours of receipt
- Independent investigation team appointed
- Due process rights of accused persons respected
- Investigation completed within 30 days
- Findings reported to CEO and Board
- Survivor support provided throughout

## 6. Disciplinary Actions
Violations result in:
- Verbal warning (minor boundary issues)
- Written warning and mandatory retraining
- Suspension pending investigation
- Termination of employment/contract
- Referral to law enforcement (criminal acts)
- Notification to relevant UN/donor agencies

## 7. Survivor Support
- Immediate medical referral
- Psychosocial support services
- Legal assistance if requested
- Protection from retaliation
- Ongoing case management

## 8. Training Requirements
- Mandatory induction training for all new personnel
- Annual refresher training for all staff
- Specialized training for managers and investigators
- Community awareness sessions in program areas
- Training attendance tracked and documented

## 9. Monitoring and Review
- Annual policy review by Board
- Quarterly PSEA compliance reporting
- External PSEA audit every 2 years
- Integration with program monitoring

Approved by:
Dr. Amina Osman, Board Chairperson
Date: December 15, 2024
'''
        }
    ])
    print("  Created: valid_psea_policy.pdf")


def create_valid_project_report():
    """5. Valid project report with indicators."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_project_report.pdf'), [
        {
            'title': 'PROJECT PROGRESS REPORT - Q1 2026',
            'body': '''Community Health Workers Scale-Up Program
Grant: Global Health Fund CHW-2025-001
Period: January - March 2026
Organization: Amani Community Development

## Executive Summary
The CHW Scale-Up Program made significant progress in Q1 2026. We trained 45 new community health workers across 3 counties, achieving 90% of quarterly target. Community health outreach reached 12,500 beneficiaries, exceeding the target of 10,000. Key challenges included supply chain delays for essential medical kits.

## Activities and Outputs
1. CHW Training Program: 45 CHWs completed 10-day training (target: 50)
2. Community Outreach: 156 outreach sessions conducted across 12 sub-counties
3. Health Screenings: 3,200 screenings conducted (malaria, nutrition, maternal health)
4. Referral System: 487 referrals to health facilities, 92% completion rate
5. Data Collection: Mobile health data system deployed in 8 health facilities

## Progress Against Indicators
Indicator 1: Number of CHWs trained
- Baseline: 120 | Target: 200 (by Dec 2026) | Q1 Actual: 165
- Progress: 82.5% of annual target

Indicator 2: Beneficiaries reached with health services
- Baseline: 8,000 | Target: 50,000 (annual) | Q1 Actual: 12,500
- Progress: 25% of annual target (on track)

Indicator 3: Under-5 malaria case detection rate
- Baseline: 34% | Target: 65% | Q1 Actual: 48%
- Progress: Improving, additional training needed

## Challenges and Mitigation
1. Supply chain delays: Medical kit delivery delayed by 3 weeks. Mitigation: Pre-ordered Q2 supplies, identified backup supplier.
2. Staff turnover: 2 CHWs dropped out. Mitigation: Reserve list activated, replacements in training.
3. Rainy season access: Some communities inaccessible in March. Mitigation: Adjusted schedule, used motorcycle transport.

## Lessons Learned
- Peer mentoring between experienced and new CHWs significantly improved retention
- Community engagement meetings before program launch improved acceptance rates
- Mobile data collection reduced reporting time by 60%

## Financial Summary
Q1 Budget: $125,000 | Q1 Actual: $118,500 | Utilization: 94.8%
Key variances: Under-spend on equipment (delayed procurement), slight over-spend on transport.

## Next Steps (Q2 2026)
- Train remaining 35 CHWs to reach annual target
- Deploy medical kits to all CHW stations
- Launch community health awareness campaign
- Conduct mid-term program evaluation
'''
        }
    ])
    print("  Created: valid_project_report.pdf")


def create_valid_strategic_plan():
    """7. Valid strategic plan."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_strategic_plan.pdf'), [
        {
            'title': 'STRATEGIC PLAN 2025-2030',
            'body': '''Amani Community Development
"Building Healthy Communities Together"

## Vision
A world where every community in East Africa has access to quality healthcare, clean water, and nutrition security.

## Mission
To strengthen community health systems, improve water and sanitation infrastructure, and build local capacity for sustainable development in underserved communities across Kenya, Somalia, and Uganda.

## Core Values
- Community Ownership: Programs designed and led by communities
- Accountability: Transparent use of resources and evidence-based decisions
- Innovation: Embracing technology and new approaches
- Equity: Reaching the most vulnerable and marginalized
- Partnership: Collaborating with governments, donors, and peer organizations

## Strategic Goal 1: Expand Community Health Services
Objective 1.1: Train 500 new community health workers by 2028
Objective 1.2: Achieve 90% coverage in target communities
Objective 1.3: Reduce under-5 mortality by 25% in program areas

## Strategic Goal 2: Strengthen WASH Infrastructure
Objective 2.1: Construct 50 community water points by 2030
Objective 2.2: Build 200 household sanitation facilities annually
Objective 2.3: Achieve open-defecation-free status in 30 communities

## Strategic Goal 3: Build Organizational Resilience
Objective 3.1: Diversify funding to reduce donor dependency below 60%
Objective 3.2: Achieve CHS certification by 2027
Objective 3.3: Implement digital M&E system across all programs
Objective 3.4: Develop succession plan for all leadership positions

## Resource Requirements
Annual budget growth from $2.5M (2025) to $5M (2030)
Staff growth from 85 to 150 full-time employees
Technology investment: $200,000 over 5 years

## Monitoring Framework
- Annual strategic plan review
- Quarterly KPI dashboard reporting
- Mid-term evaluation (2027)
- End-term evaluation (2030)
'''
        }
    ])
    print("  Created: valid_strategic_plan.pdf")


def create_valid_org_chart():
    """8. Valid organizational chart."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_org_chart.pdf'), [
        {
            'title': 'ORGANIZATIONAL STRUCTURE',
            'body': '''Amani Community Development - March 2026

BOARD OF DIRECTORS (7 members)
Dr. Amina Osman (Chair) | James Kariuki (Secretary) | Grace Wanjiku (Treasurer)
4 Independent Directors

|
EXECUTIVE DIRECTOR
Fatima Hassan

|
+--- FINANCE & ADMIN DIRECTOR --- Finance Manager --- 2 Finance Officers
|                                  Admin Manager --- 3 Admin Staff
|                                  Procurement Officer
|
+--- PROGRAMS DIRECTOR --- Health Program Manager --- 4 Program Officers
|                          WASH Program Manager --- 3 Program Officers
|                          Nutrition Coordinator --- 2 Program Officers
|
+--- M&E DIRECTOR --- M&E Manager --- 2 M&E Officers
|                     Data Analyst
|                     Learning Coordinator
|
+--- HR & SAFEGUARDING MANAGER --- HR Officer
                                    Safeguarding Focal Point

Total Staff: 85 (65 full-time, 20 part-time/consultants)
Field offices: Nairobi (HQ), Mombasa, Garissa, Kisumu
'''
        }
    ])
    print("  Created: valid_org_chart.pdf")


def create_valid_cv():
    """9. Valid CV for program manager."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_cv_program_manager.pdf'), [
        {
            'title': 'CURRICULUM VITAE',
            'body': '''Name: Dr. James Mwangi Ochieng
Position: Health Program Manager
Nationality: Kenyan
Languages: English (fluent), Swahili (native), Somali (basic)

## Professional Summary
Public health professional with 12 years of experience in community health programming, project management, and capacity building in East Africa. Expertise in CHW training, maternal and child health, and health systems strengthening.

## Education
- Master of Public Health (MPH), University of Nairobi, 2014
- Bachelor of Medicine (MBChB), Kenyatta University, 2011
- Certificate in Project Management (PMP), PMI, 2018

## Professional Experience

Health Program Manager, Amani Community Development (2019-Present)
- Led CHW Scale-Up Program reaching 50,000 beneficiaries across 3 counties
- Managed annual budget of $800,000 with 98% utilization rate
- Supervised team of 4 program officers and 120 community health workers
- Developed M&E framework achieving 95% data completeness
- Published 2 peer-reviewed articles on community health

Senior Health Officer, UNICEF Kenya (2015-2019)
- Supported Ministry of Health in CHW policy development
- Coordinated national immunization campaigns reaching 2.5M children
- Conducted health facility assessments in 5 counties

Medical Officer, Kenyatta National Hospital (2011-2015)
- Provided clinical care in pediatrics and maternal health departments
- Participated in research on malaria case management

## Key Skills
- Program design and management
- Budget management and donor reporting
- Team leadership and mentoring
- Data analysis (STATA, SPSS, Excel)
- Report writing and presentation
- Stakeholder engagement

## References
Available upon request
'''
        }
    ])
    print("  Created: valid_cv_program_manager.pdf")


def create_valid_reference_letter():
    """10. Valid donor reference letter."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'valid_reference_letter.pdf'), [
        {
            'title': 'LETTER OF REFERENCE',
            'body': '''USAID Kenya and East Africa Mission
United States Agency for International Development
P.O. Box 629, Village Market 00621
Nairobi, Kenya

Date: January 15, 2026

To Whom It May Concern:

RE: Reference for Amani Community Development (NGO/2012/3847)

I am writing to provide a reference for Amani Community Development, with whom USAID has maintained a productive partnership since 2018.

Amani has been a sub-grantee under our Community Health Strengthening Activity (CHSA), receiving a total of $1.2 million over three years (2021-2024). Throughout this period, they have demonstrated:

1. Strong Financial Management: All financial reports submitted on time with minimal questioned costs (less than 0.5% of total expenditure). Successfully passed annual financial audits with unqualified opinions.

2. Effective Program Delivery: Exceeded program targets by an average of 15%. Their community health worker model has been recognized as a best practice by the Kenyan Ministry of Health.

3. Compliance: Full compliance with USAID regulations including environmental compliance, PSEA requirements, and Trafficking in Persons policy.

4. Reporting Quality: High-quality narrative and financial reports. Responsive to feedback and revision requests.

5. Organizational Capacity: Well-structured governance with an active board. Clear policies and procedures. Strong safeguarding framework.

We would not hesitate to partner with Amani again and consider them a reliable implementing partner in the East African humanitarian and development context.

For any queries, please contact:

Dr. Sarah Johnson
Agreement Officer Representative
USAID Kenya
Tel: +254-20-XXX-XXXX
Email: sjohnson@usaid.gov
'''
        }
    ])
    print("  Created: valid_reference_letter.pdf")


def create_valid_budget_xlsx():
    """6. Valid budget spreadsheet."""
    if not HAS_OPENPYXL:
        print("  SKIPPED: valid_budget_detail.xlsx (openpyxl not installed)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Detail"

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    currency_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Amani Community Development - Project Budget Detail'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Community Health Workers Scale-Up Program | Grant Period: Jan 2025 - Dec 2026'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)

    # Headers
    headers = ['Category', 'Line Item', 'Budget (USD)', 'Actual (USD)', 'Variance (USD)', 'Utilization %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data
    data = [
        ('Personnel', 'Program Manager (1)', 48000, 48000),
        ('Personnel', 'Program Officers (4)', 96000, 94500),
        ('Personnel', 'CHW Supervisors (6)', 54000, 52800),
        ('Personnel', 'M&E Officer (1)', 24000, 24000),
        ('Personnel', 'Finance Officer (1)', 22000, 22000),
        ('Travel', 'Field travel - vehicles', 36000, 33500),
        ('Travel', 'Per diem - field staff', 18000, 17200),
        ('Travel', 'International travel', 8000, 4500),
        ('Equipment', 'Medical kits (120)', 60000, 58000),
        ('Equipment', 'Smartphones for CHWs', 24000, 23400),
        ('Equipment', 'Weighing scales (50)', 5000, 4800),
        ('Supplies', 'Training materials', 12000, 11500),
        ('Supplies', 'IEC materials', 8000, 7800),
        ('Supplies', 'Office supplies', 4000, 3800),
        ('Overhead', 'Office rent', 18000, 18000),
        ('Overhead', 'Utilities', 6000, 5800),
        ('Overhead', 'Insurance', 4000, 4000),
        ('Overhead', 'Audit fees', 8000, 8000),
        ('Overhead', 'Communication', 5000, 4600),
        ('Monitoring', 'Baseline survey', 15000, 14500),
        ('Monitoring', 'Mid-term evaluation', 10000, 0),
        ('Monitoring', 'Data management system', 8000, 7500),
        ('Contingency', 'Contingency (5%)', 25000, 3200),
    ]

    for i, (cat, item, budget, actual) in enumerate(data, 5):
        ws.cell(row=i, column=1, value=cat).border = thin_border
        ws.cell(row=i, column=2, value=item).border = thin_border
        c3 = ws.cell(row=i, column=3, value=budget)
        c3.number_format = currency_fmt
        c3.border = thin_border
        c4 = ws.cell(row=i, column=4, value=actual)
        c4.number_format = currency_fmt
        c4.border = thin_border
        c5 = ws.cell(row=i, column=5)
        c5.value = budget - actual
        c5.number_format = currency_fmt
        c5.border = thin_border
        c6 = ws.cell(row=i, column=6)
        c6.value = actual / budget if budget > 0 else 0
        c6.number_format = pct_fmt
        c6.border = thin_border

    # Totals row
    total_row = 5 + len(data)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=2).border = thin_border
    for col in [3, 4, 5]:
        c = ws.cell(row=total_row, column=col)
        c.value = f'=SUM({chr(64+col)}5:{chr(64+col)}{total_row-1})'
        c.number_format = currency_fmt
        c.font = Font(bold=True)
        c.border = thin_border
    c6 = ws.cell(row=total_row, column=6)
    c6.value = f'=D{total_row}/C{total_row}'
    c6.number_format = pct_fmt
    c6.font = Font(bold=True)
    c6.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    wb.save(os.path.join(TEST_DIR, 'valid_budget_detail.xlsx'))
    print("  Created: valid_budget_detail.xlsx")


# ═══════════════════════════════════════════════════════════════════
# NEGATIVE TEST FILES
# ═══════════════════════════════════════════════════════════════════

def create_invalid_empty_pdf():
    """Empty/minimal PDF with no real content."""
    # Valid PDF structure but with blank page
    if HAS_FPDF:
        pdf = FPDF()
        pdf.add_page()
        # Intentionally no content
        pdf.output(os.path.join(TEST_DIR, 'invalid_empty.pdf'))
    else:
        raw = (
            b'%PDF-1.4\n'
            b'1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
            b'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n'
            b'3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj\n'
            b'xref\n0 4\n'
            b'0000000000 65535 f \n'
            b'0000000009 00000 n \n'
            b'0000000058 00000 n \n'
            b'0000000115 00000 n \n'
            b'trailer<</Size 4/Root 1 0 R>>\n'
            b'startxref\n175\n%%EOF\n'
        )
        with open(os.path.join(TEST_DIR, 'invalid_empty.pdf'), 'wb') as f:
            f.write(raw)
    print("  Created: invalid_empty.pdf")


def create_invalid_wrong_extension():
    """Text file with .pdf extension."""
    path = os.path.join(TEST_DIR, 'invalid_wrong_extension.pdf')
    with open(path, 'w') as f:
        f.write("This is a plain text file disguised as a PDF.\n")
        f.write("It does not contain valid PDF content.\n")
        f.write("The system should detect this via magic byte validation.\n")
    print("  Created: invalid_wrong_extension.pdf")


def create_invalid_oversized():
    """File that exceeds 16MB limit."""
    path = os.path.join(TEST_DIR, 'invalid_oversized.pdf')
    if HAS_FPDF:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Helvetica', '', 10)
        # Generate enough content to exceed 16MB
        # Each line is ~100 bytes, need ~170,000 lines
        filler = "This is filler content to create an oversized test file for upload validation testing. " * 5
        for i in range(3500):
            pdf.add_page()
            for j in range(40):
                pdf.cell(0, 5, f"Page {i} Line {j}: {filler}", new_x='LMARGIN', new_y='NEXT')
        pdf.output(path)
    else:
        # Create a large file that starts with PDF header
        with open(path, 'wb') as f:
            f.write(b'%PDF-1.4\n')
            # Write 17MB of data
            chunk = b'X' * (1024 * 1024)  # 1MB chunks
            for _ in range(17):
                f.write(chunk)
            f.write(b'\n%%EOF\n')
    size_mb = os.path.getsize(path) / (1024 * 1024)
    print(f"  Created: invalid_oversized.pdf ({size_mb:.1f} MB)")


def create_invalid_executable():
    """Executable file that should be rejected."""
    path = os.path.join(TEST_DIR, 'invalid_malware_test.exe')
    with open(path, 'wb') as f:
        # MZ header (Windows PE signature) - clearly an executable
        f.write(b'MZ')
        f.write(b'\x90' * 58)  # DOS header padding
        f.write(b'This is a test executable file. Not actual malware.\n')
        f.write(b'The upload system should reject this file type.\n')
    print("  Created: invalid_malware_test.exe")


def create_invalid_financial_incomplete():
    """Financial report missing critical sections."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'invalid_financial_incomplete.pdf'), [
        {
            'title': 'FINANCIAL SUMMARY - PARTIAL',
            'body': '''Organization: Test NGO
Period: Unknown

Some expenses were incurred.
Total: approximately $50,000

Note: This report is intentionally incomplete for testing purposes.
- No revenue breakdown provided
- No budget vs. actual comparison
- No auditor opinion
- No cash flow statement
- No internal controls section
- No donor compliance information
- No supporting documentation referenced
'''
        }
    ])
    print("  Created: invalid_financial_incomplete.pdf")


def create_expired_registration():
    """Registration certificate with expired date."""
    make_minimal_pdf(os.path.join(TEST_DIR, 'expired_registration.pdf'), [
        {
            'title': 'CERTIFICATE OF REGISTRATION (EXPIRED)',
            'body': '''Republic of Kenya
NGO Coordination Board

CERTIFICATE OF REGISTRATION

Organization Name: Old Charity Foundation
Registration Number: NGO/2010/1234
Date of Registration: June 30, 2010
Date of Expiry: June 30, 2023     *** EXPIRED ***

Status: EXPIRED - RENEWAL REQUIRED

This certificate has expired and is no longer valid.
The organization must apply for renewal before conducting any activities.

Warning: Operating with an expired registration may result in
penalties under the NGO Coordination Act.

Last Valid Date: June 30, 2023
Current Date: March 2026
Days Expired: 974 days
'''
        }
    ])
    print("  Created: expired_registration.pdf")


def create_malicious_html():
    """HTML file - should be rejected by file type filter."""
    path = os.path.join(TEST_DIR, 'malicious_script.html')
    with open(path, 'w') as f:
        f.write('<!DOCTYPE html>\n<html>\n<head><title>Test</title></head>\n')
        f.write('<body>\n<h1>This is an HTML file</h1>\n')
        f.write('<p>The upload system should reject HTML files.</p>\n')
        f.write('<p>Only PDF, DOCX, XLSX, CSV, PNG, JPG, TXT are allowed.</p>\n')
        f.write('</body>\n</html>\n')
    print("  Created: malicious_script.html")


def create_valid_csv_data():
    """Valid CSV data file for testing."""
    path = os.path.join(TEST_DIR, 'valid_beneficiary_data.csv')
    with open(path, 'w') as f:
        f.write('Date,Location,Beneficiary_Type,Count,Service_Provided,Notes\n')
        f.write('2026-01-15,Kibera,Women,45,Health Screening,Completed successfully\n')
        f.write('2026-01-15,Kibera,Children U5,78,Nutrition Assessment,3 referred for SAM treatment\n')
        f.write('2026-01-22,Mathare,Women,52,Antenatal Care,12 high-risk pregnancies identified\n')
        f.write('2026-01-22,Mathare,Men,23,Health Education,HIV testing counseling\n')
        f.write('2026-02-01,Dandora,Children U5,91,Immunization,Measles and polio vaccines\n')
        f.write('2026-02-01,Dandora,Elderly,34,Chronic Disease Screening,8 diabetes cases found\n')
        f.write('2026-02-15,Korogocho,Women,67,Family Planning,Contraceptive counseling and provision\n')
        f.write('2026-02-15,Korogocho,Youth,42,Reproductive Health,Peer education session\n')
        f.write('2026-03-01,Kibera,Women,55,Health Screening,Follow-up screenings\n')
        f.write('2026-03-01,Kibera,Children U5,83,Growth Monitoring,5 moderate malnutrition cases\n')
    print("  Created: valid_beneficiary_data.csv")


def create_valid_txt_report():
    """Valid text file for testing."""
    path = os.path.join(TEST_DIR, 'valid_meeting_minutes.txt')
    with open(path, 'w') as f:
        f.write("BOARD MEETING MINUTES\n")
        f.write("=" * 50 + "\n")
        f.write("Date: February 15, 2026\n")
        f.write("Location: Amani HQ, Nairobi\n")
        f.write("Present: Dr. Amina Osman (Chair), James Kariuki, Grace Wanjiku, Fatima Hassan\n")
        f.write("Apologies: 2 board members\n\n")
        f.write("AGENDA:\n")
        f.write("1. Approval of previous minutes\n")
        f.write("2. Financial report Q4 2025\n")
        f.write("3. Strategic plan progress review\n")
        f.write("4. New grant opportunities\n")
        f.write("5. AOB\n\n")
        f.write("RESOLUTIONS:\n")
        f.write("1. Minutes of November 2025 meeting approved unanimously\n")
        f.write("2. Q4 financial report noted. Budget utilization at 94.2%.\n")
        f.write("3. Strategic plan on track. 3 of 5 year-1 targets achieved.\n")
        f.write("4. Board approved application to Global Health Fund CHW program.\n")
        f.write("5. Next meeting: May 15, 2026\n\n")
        f.write("Minutes prepared by: James Kariuki, Secretary\n")
    print("  Created: valid_meeting_minutes.txt")


# ═══════════════════════════════════════════════════════════════════
# MAIN: Generate all files and create ZIP
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  Kuja Grant v3.0 - Test File Generator")
    print("=" * 60)
    print()

    print("Creating POSITIVE test files...")
    create_valid_financial_report()
    create_valid_registration_cert()
    create_valid_audit_report()
    create_valid_psea_policy()
    create_valid_project_report()
    create_valid_budget_xlsx()
    create_valid_strategic_plan()
    create_valid_org_chart()
    create_valid_cv()
    create_valid_reference_letter()
    create_valid_csv_data()
    create_valid_txt_report()

    print()
    print("Creating NEGATIVE test files...")
    create_invalid_empty_pdf()
    create_invalid_wrong_extension()
    create_invalid_oversized()
    create_invalid_executable()
    create_invalid_financial_incomplete()
    create_expired_registration()
    create_malicious_html()

    print()
    print("Packaging into ZIP...")
    zip_path = os.path.join(DOCS_DIR, 'kuja-test-files-v3.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname in sorted(os.listdir(TEST_DIR)):
            fpath = os.path.join(TEST_DIR, fname)
            if os.path.isfile(fpath):
                zf.write(fpath, f'kuja-test-files-v3/{fname}')
                size = os.path.getsize(fpath)
                print(f"  Added: {fname} ({size:,} bytes)")

    zip_size = os.path.getsize(zip_path)
    file_count = len([f for f in os.listdir(TEST_DIR) if os.path.isfile(os.path.join(TEST_DIR, f))])
    print()
    print("=" * 60)
    print(f"  COMPLETE: {file_count} test files created")
    print(f"  ZIP: {zip_path}")
    print(f"  ZIP Size: {zip_size:,} bytes ({zip_size/1024:.1f} KB)")
    print("=" * 60)


if __name__ == '__main__':
    main()
