"""Generate an Excel workbook of the Kuja v5.0 test cases.

Output: Kuja_Grant_v5.0_Test_Cases.xlsx

Why an xlsx (in addition to the docx):
  - One tab per module (32 tabs covering 306 cases)
  - Pass/Fail + Tester column on every row so the team can track runs in-place
  - Summary tab with category counts + live pass/fail rollup formulas
  - Color-coded priority cells (P1 red, P2 amber, P3 grey)
  - Conditional formatting on Status cells (PASS green, FAIL red, BLOCKED amber)
  - Frozen header row + sensible column widths so steps/expected wrap cleanly

Source of truth: imports the `test_cases` list from generate_test_cases_doc.py
so both generators stay in lock-step. Add a test case once, it appears in
both the docx and the xlsx automatically.
"""

import os
import re
import sys

# Make sibling generator importable when run from anywhere
THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, THIS_DIR)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# Import the canonical test_cases list (built at module import; the
# docx-write side effect lives under __main__, so this import is clean).
from generate_test_cases_doc import test_cases


OUT_PATH = os.path.join(THIS_DIR, "Kuja_Grant_v5.0_Test_Cases.xlsx")


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
KUJA_NAVY = "1A237E"
KUJA_NAVY_LIGHT = "C5CAE9"
KUJA_CLAY = "C2410C"
KUJA_GROW = "1B5E20"
KUJA_FLAG = "B71C1C"
KUJA_AMBER = "EF6C00"
KUJA_SAND = "FFF8E1"
KUJA_GREY = "ECEFF1"

PRIORITY_FILL = {
    "P1": PatternFill("solid", fgColor="FFCDD2"),  # light red
    "P2": PatternFill("solid", fgColor="FFE0B2"),  # light amber
    "P3": PatternFill("solid", fgColor="E0E0E0"),  # light grey
}

HEADER_FILL = PatternFill("solid", fgColor=KUJA_NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
BODY_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

STATUS_OPTIONS = ["", "PASS", "FAIL", "BLOCKED", "N/A", "IN PROGRESS"]


# ---------------------------------------------------------------------------
# Column layout (per test-case tab)
# ---------------------------------------------------------------------------
COLUMNS = [
    ("ID",                  10),
    ("Test Case",           28),
    ("Priority",            11),
    ("Type",                14),
    ("Requirement",         16),
    ("Environment / Device", 16),
    ("Data Dependency",     22),
    ("Cleanup",             18),
    ("Prerequisites",       30),
    ("Steps",               50),
    ("Test Data",           24),
    ("Expected Result",     50),
    ("Pass Criteria",       28),
    ("Automation",          16),
    ("Status",              14),
    ("Tester",              12),
    ("Run Date",            12),
    ("Build Verified",      16),
    ("Comments / Defect ID", 32),
]

# Column indexes used by helpers (1-based).
COL_PRIORITY = 3
COL_TYPE = 4
COL_AUTOMATION = 14
COL_STATUS = 15


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: <=31 chars, no \\ / ? * [ ]"""
    cleaned = re.sub(r"[\\/?*\[\]:]", " ", name).strip()
    return cleaned[:31] if len(cleaned) > 31 else cleaned


def _priority_short(p: str) -> str:
    if "P1" in p:
        return "P1"
    if "P2" in p:
        return "P2"
    if "P3" in p:
        return "P3"
    return ""


def _write_header(ws):
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _coerce_text(v) -> str:
    """Some test-case fields use Python string-concat across multiple
    parenthesised lines. A stray comma turns the field into a tuple. The
    docx generator silently handles either; we have to be explicit
    because Excel cells refuse non-scalar values. Join tuples/lists with
    newlines, coerce everything else with str()."""
    if v is None:
        return ""
    if isinstance(v, (tuple, list)):
        return "\n".join(_coerce_text(x) for x in v if x is not None)
    return str(v)


# Inferred-default helpers — the team's 2026-05-17 feedback called for
# extra tracking columns (Type, Automation Candidate, Environment, Data
# Dependency, Cleanup, Build Verified). Existing test cases don't carry
# explicit values for these so we infer sensible defaults from category
# and id ranges. New tests can set the fields explicitly and override.

def _infer_type(tc: dict) -> str:
    """Returns one of: happy_path, negative, edge, security, privacy,
    accessibility, performance, i18n, analytics, e2e."""
    if tc.get("type"):
        return tc["type"]
    cat = (tc.get("category") or "").lower()
    name = (tc.get("name") or "").lower()
    if "accessibility" in cat:
        return "accessibility"
    if "performance" in cat or "slo" in cat:
        return "performance"
    if "privacy" in cat or "leakage" in cat:
        return "privacy"
    if "cross-browser" in cat or "device matrix" in cat:
        return "edge"
    if "integrity" in cat or "workflow" in cat:
        return "edge"
    if "i18n" in cat or "rtl" in cat:
        return "i18n"
    if "metrics" in cat or "feedback" in cat:
        return "analytics"
    if "e2e" in cat or "workflows" in cat:
        return "e2e"
    if "edge" in cat:
        return "edge"
    if "negative" in cat:
        return "negative"
    if "forbidden" in name or "rejected" in name or "blocked" in name or "denied" in name:
        return "negative"
    if "xss" in name or "injection" in name or "auth" in name and "begin" not in name:
        return "security"
    return "happy_path"


def _infer_automation(tc: dict) -> str:
    """Returns one of: manual, api, browser, already_automated, not_worth_automating."""
    if tc.get("automation"):
        return tc["automation"]
    cat = (tc.get("category") or "").lower()
    if "accessibility" in cat or "cross-browser" in cat or "device matrix" in cat:
        return "browser"
    if "privacy" in cat or "metrics" in cat or "integrity" in cat:
        return "api"
    if "performance" in cat or "slo" in cat:
        return "browser"
    return "manual"


def _infer_environment(tc: dict) -> str:
    if tc.get("environment"):
        return tc["environment"]
    name = (tc.get("name") or "").lower()
    cat = (tc.get("category") or "").lower()
    if "ios" in name or "iphone" in name or "android" in name or "mobile" in name:
        return "mobile"
    if "safari" in name:
        return "Safari"
    if "firefox" in name:
        return "Firefox"
    if "edge" in cat:  # category "Edge" — not Edge browser
        return "any"
    if "rtl" in cat or "arabic" in name or "i18n" in cat:
        return "RTL-capable browser"
    return "desktop"


def _infer_data_dep(tc: dict) -> str:
    if tc.get("data_dep"):
        return tc["data_dep"]
    pre = (tc.get("prereqs") or "").lower()
    # Extract the most important phrase from prereqs as a quick summary
    if "fresh" in pre:
        return "fresh account"
    if "no credentials" in pre or "no enrolled" in pre:
        return "user with no WebAuthn credentials"
    if "admin" in pre:
        return "admin account"
    if "reviewer" in pre:
        return "reviewer with assignments"
    if "draft application" in pre:
        return "NGO with draft application"
    if "submitted" in pre and "application" in pre:
        return "submitted application"
    if "awarded" in pre:
        return "awarded application"
    if "donor with" in pre:
        return "donor + portfolio"
    if "two ngo" in pre or "two donor" in pre or "two user" in pre:
        return "two accounts (isolation test)"
    return "default seed"


def _infer_cleanup(tc: dict) -> str:
    if tc.get("cleanup"):
        return tc["cleanup"]
    name = (tc.get("name") or "").lower()
    cat = (tc.get("category") or "").lower()
    if "lockout" in name:
        return "clear-lockout endpoint"
    if "feedback" in cat or "broadcast" in name or "decision" in name:
        return "leaves data — auto-rolled into metrics"
    if "create" in name or "draft" in name or "submit" in name:
        return "delete created entity or auto-expire"
    if "performance" in cat or "concurrency" in name:
        return "none"
    if "accessibility" in cat or "cross-browser" in cat:
        return "none"
    return "none"


TYPE_FILL = {
    "happy_path":     PatternFill("solid", fgColor="E1F5FE"),  # light blue
    "negative":       PatternFill("solid", fgColor="FFCDD2"),  # light red
    "edge":           PatternFill("solid", fgColor="FFE0B2"),  # light amber
    "security":       PatternFill("solid", fgColor="FFCCBC"),  # peach
    "privacy":        PatternFill("solid", fgColor="D7CCC8"),  # warm grey
    "accessibility":  PatternFill("solid", fgColor="C8E6C9"),  # light green
    "performance":    PatternFill("solid", fgColor="B3E5FC"),  # cyan
    "i18n":           PatternFill("solid", fgColor="E1BEE7"),  # lavender
    "analytics":      PatternFill("solid", fgColor="D1C4E9"),  # purple
    "e2e":            PatternFill("solid", fgColor="FFF9C4"),  # pale yellow
}


def _write_test_case(ws, row_idx: int, tc: dict):
    priority_short = _priority_short(tc.get("priority", ""))
    type_value = _infer_type(tc)
    automation_value = _infer_automation(tc)
    environment_value = _infer_environment(tc)
    data_dep_value = _infer_data_dep(tc)
    cleanup_value = _infer_cleanup(tc)

    row_values = [
        _coerce_text(tc.get("id")),
        _coerce_text(tc.get("name")),
        _coerce_text(tc.get("priority")),
        type_value,
        _coerce_text(tc.get("requirement")),
        environment_value,
        data_dep_value,
        cleanup_value,
        _coerce_text(tc.get("prereqs")),
        _coerce_text(tc.get("steps")),
        _coerce_text(tc.get("data")),
        _coerce_text(tc.get("expected")),
        _coerce_text(tc.get("criteria")),
        automation_value,
        "",  # Status (tester)
        "",  # Tester
        "",  # Run Date
        "",  # Build Verified
        "",  # Comments
    ]
    centered_cols = {COL_PRIORITY, COL_TYPE, 6, 8, COL_AUTOMATION,
                     COL_STATUS, COL_STATUS + 1, COL_STATUS + 2, COL_STATUS + 3}
    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = BODY_ALIGN_CENTER if col_idx in centered_cols else BODY_ALIGN

    # Color the priority cell
    priority_cell = ws.cell(row=row_idx, column=COL_PRIORITY)
    if priority_short in PRIORITY_FILL:
        priority_cell.fill = PRIORITY_FILL[priority_short]
        priority_cell.font = Font(name="Calibri", size=10, bold=True)

    # Color the type cell
    type_cell = ws.cell(row=row_idx, column=COL_TYPE)
    if type_value in TYPE_FILL:
        type_cell.fill = TYPE_FILL[type_value]
        type_cell.font = Font(name="Calibri", size=10, bold=True)


def _add_status_validation_and_formatting(ws, last_row: int):
    """Dropdown on Status column + conditional fill for PASS/FAIL/BLOCKED."""
    status_col_letter = get_column_letter(COL_STATUS)
    rng = f"{status_col_letter}2:{status_col_letter}{last_row}"

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(s for s in STATUS_OPTIONS if s)}"',
        allow_blank=True,
        showDropDown=False,  # showDropDown=False → arrow IS shown (openpyxl quirk)
    )
    dv.add(rng)
    ws.add_data_validation(dv)

    # Conditional formatting for status colors
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"PASS"'],
        fill=PatternFill("solid", fgColor="C8E6C9"),  # light green
        font=Font(name="Calibri", size=10, bold=True, color=KUJA_GROW),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"FAIL"'],
        fill=PatternFill("solid", fgColor="FFCDD2"),  # light red
        font=Font(name="Calibri", size=10, bold=True, color=KUJA_FLAG),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"BLOCKED"'],
        fill=PatternFill("solid", fgColor="FFECB3"),  # light amber
        font=Font(name="Calibri", size=10, bold=True, color=KUJA_AMBER),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"IN PROGRESS"'],
        fill=PatternFill("solid", fgColor="BBDEFB"),  # light blue
        font=Font(name="Calibri", size=10, bold=True, color="0D47A1"),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"N/A"'],
        fill=PatternFill("solid", fgColor="EEEEEE"),
        font=Font(name="Calibri", size=10, italic=True, color="616161"),
    ))


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook():
    # Group test cases by category, preserving the order they appear in the source list
    seen = []
    by_category: dict[str, list[dict]] = {}
    for tc in test_cases:
        cat = tc.get("category", "Uncategorised")
        if cat not in by_category:
            seen.append(cat)
            by_category[cat] = []
        by_category[cat].append(tc)

    wb = Workbook()

    # ── README sheet ──
    readme = wb.active
    readme.title = "README"
    readme["A1"] = "Kuja Grant — Test Cases v5.0"
    readme["A1"].font = Font(name="Calibri", size=20, bold=True, color=KUJA_NAVY)
    readme["A2"] = f"{len(test_cases)} test cases across {len(seen)} modules"
    readme["A2"].font = Font(name="Calibri", size=12, italic=True, color="424242")
    readme["A2"].alignment = Alignment(horizontal="left")

    lines = [
        "",
        "HOW TO USE THIS WORKBOOK",
        "",
        "  1. One sheet per module/category (see tabs at the bottom).",
        "  2. Pick a sheet, work through test cases row by row.",
        "  3. On each row, fill in:",
        "       • Status: pick from dropdown (PASS / FAIL / BLOCKED / N/A / IN PROGRESS)",
        "       • Tester: your name or initials",
        "       • Run Date: when you executed the test",
        "       • Comments / Defect ID: notes, screenshots, JIRA/issue link",
        "",
        "  4. The Summary sheet auto-rolls up PASS/FAIL counts via live formulas.",
        "  5. Status cells are colour-coded automatically:",
        "       • PASS = green",
        "       • FAIL = red (this is your defect)",
        "       • BLOCKED = amber (you couldn't run — note why in Comments)",
        "       • IN PROGRESS = blue",
        "       • N/A = grey (not applicable in your environment)",
        "",
        "  6. Priority column is colour-coded for triage:",
        "       • P1 Critical = red — must pass before deploy",
        "       • P2 High     = amber — strongly recommended",
        "       • P3 Medium   = grey — nice to have",
        "",
        "REGENERATION",
        "",
        "  This file is generated from `generate_test_cases_xlsx.py`, which",
        "  imports the canonical test_cases list from generate_test_cases_doc.py.",
        "  Run `py -3 docs/generate_test_cases_xlsx.py` to produce a fresh copy.",
        "",
        "  IMPORTANT: re-running OVERWRITES the file. Save your in-progress",
        "  Status/Tester/Comments data to a copy first.",
        "",
        f"  Generated against commit at build time.",
    ]
    for i, line in enumerate(lines, start=3):
        cell = readme.cell(row=i, column=1, value=line)
        if line and line == line.upper() and not line.startswith(" "):
            cell.font = Font(name="Calibri", size=12, bold=True, color=KUJA_CLAY)
        else:
            cell.font = Font(name="Calibri", size=11, color="212121")
    readme.column_dimensions["A"].width = 110

    # ── Summary sheet (will be populated after we know sheet names) ──
    summary = wb.create_sheet("Summary")

    # ── One sheet per category ──
    sheet_meta = []  # list of (sheet_name, category, first_data_row, last_data_row)
    for cat in seen:
        sheet_name = _safe_sheet_name(cat)
        # Disambiguate if collision after truncation
        base = sheet_name
        suffix = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base[:28]}_{suffix}"
            suffix += 1

        ws = wb.create_sheet(sheet_name)
        _write_header(ws)

        first_row = 2
        for offset, tc in enumerate(by_category[cat]):
            _write_test_case(ws, first_row + offset, tc)
            ws.row_dimensions[first_row + offset].height = 110

        last_row = first_row + len(by_category[cat]) - 1
        _add_status_validation_and_formatting(ws, last_row)
        sheet_meta.append((sheet_name, cat, first_row, last_row))

    # ── Populate Summary sheet ──
    summary_headers = ["Module", "Tab Link", "Total", "P1", "P2", "P3",
                       "PASS", "FAIL", "BLOCKED", "Pending", "Pass Rate"]
    for col_idx, h in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    widths = [32, 18, 8, 6, 6, 6, 8, 8, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(i)].width = w
    summary.row_dimensions[1].height = 26
    summary.freeze_panes = "A2"

    grand_p1 = grand_p2 = grand_p3 = 0
    for row_offset, (sheet_name, cat, first_row, last_row) in enumerate(sheet_meta, start=2):
        cat_cases = by_category[cat]
        total = len(cat_cases)
        p1 = sum(1 for tc in cat_cases if "P1" in tc["priority"])
        p2 = sum(1 for tc in cat_cases if "P2" in tc["priority"])
        p3 = sum(1 for tc in cat_cases if "P3" in tc["priority"])
        grand_p1 += p1
        grand_p2 += p2
        grand_p3 += p3

        # Module name + a clickable link to the tab
        summary.cell(row=row_offset, column=1, value=cat).font = Font(name="Calibri", size=10, bold=True)
        link_cell = summary.cell(row=row_offset, column=2, value="Open ›")
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(name="Calibri", size=10, color="1A237E", underline="single")
        link_cell.alignment = BODY_ALIGN_CENTER

        summary.cell(row=row_offset, column=3, value=total).alignment = BODY_ALIGN_CENTER
        summary.cell(row=row_offset, column=4, value=p1).alignment = BODY_ALIGN_CENTER
        summary.cell(row=row_offset, column=5, value=p2).alignment = BODY_ALIGN_CENTER
        summary.cell(row=row_offset, column=6, value=p3).alignment = BODY_ALIGN_CENTER

        # Live formulas counting Status cells in the source tab.
        # Status column moved from J → O (column 15) when the Type +
        # Automation + Env + Data Dep + Cleanup + Build Verified columns
        # were added in the 2026-05-17 redline pass.
        status_letter = get_column_letter(COL_STATUS)
        rng = f"'{sheet_name}'!{status_letter}{first_row}:{status_letter}{last_row}"
        summary.cell(row=row_offset, column=7, value=f'=COUNTIF({rng},"PASS")').alignment = BODY_ALIGN_CENTER
        summary.cell(row=row_offset, column=8, value=f'=COUNTIF({rng},"FAIL")').alignment = BODY_ALIGN_CENTER
        summary.cell(row=row_offset, column=9, value=f'=COUNTIF({rng},"BLOCKED")').alignment = BODY_ALIGN_CENTER
        # Pending = total - (PASS + FAIL + BLOCKED + N/A)
        pending_formula = (
            f'={total}-COUNTIF({rng},"PASS")-COUNTIF({rng},"FAIL")'
            f'-COUNTIF({rng},"BLOCKED")-COUNTIF({rng},"N/A")'
        )
        summary.cell(row=row_offset, column=10, value=pending_formula).alignment = BODY_ALIGN_CENTER
        # Pass rate = PASS / (PASS + FAIL)  (excludes blocked/pending)
        pr_formula = (
            f'=IFERROR(COUNTIF({rng},"PASS")/'
            f'(COUNTIF({rng},"PASS")+COUNTIF({rng},"FAIL")),"")'
        )
        pr_cell = summary.cell(row=row_offset, column=11, value=pr_formula)
        pr_cell.number_format = "0.0%"
        pr_cell.alignment = BODY_ALIGN_CENTER

        # Borders for the row
        for c in range(1, 12):
            summary.cell(row=row_offset, column=c).border = THIN_BORDER

    # Grand-total row
    grand_row = 2 + len(sheet_meta)
    summary.cell(row=grand_row, column=1, value="GRAND TOTAL").font = Font(
        name="Calibri", size=11, bold=True, color=KUJA_NAVY
    )
    summary.cell(row=grand_row, column=3, value=len(test_cases))
    summary.cell(row=grand_row, column=4, value=grand_p1)
    summary.cell(row=grand_row, column=5, value=grand_p2)
    summary.cell(row=grand_row, column=6, value=grand_p3)
    # Aggregate PASS/FAIL by summing the per-row formula results
    summary.cell(row=grand_row, column=7, value=f"=SUM(G2:G{grand_row - 1})")
    summary.cell(row=grand_row, column=8, value=f"=SUM(H2:H{grand_row - 1})")
    summary.cell(row=grand_row, column=9, value=f"=SUM(I2:I{grand_row - 1})")
    summary.cell(row=grand_row, column=10, value=f"=SUM(J2:J{grand_row - 1})")
    summary.cell(
        row=grand_row, column=11,
        value=f"=IFERROR(G{grand_row}/(G{grand_row}+H{grand_row}),\"\")",
    ).number_format = "0.0%"
    for c in range(1, 12):
        cell = summary.cell(row=grand_row, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True, color=KUJA_NAVY)
        cell.fill = PatternFill("solid", fgColor=KUJA_NAVY_LIGHT)
        cell.border = THIN_BORDER
        if c >= 3:
            cell.alignment = BODY_ALIGN_CENTER

    summary.row_dimensions[grand_row].height = 28

    wb.save(OUT_PATH)
    return len(test_cases), len(sheet_meta)


if __name__ == "__main__":
    total, modules = build_workbook()
    print(f"Saved {OUT_PATH}")
    print(f"  {total} test cases across {modules} module tabs")
    print(f"  Size: {os.path.getsize(OUT_PATH):,} bytes")
